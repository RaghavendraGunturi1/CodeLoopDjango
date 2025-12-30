import requests
import logging
import json # <-- ADD json import
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.utils import timezone
from django.contrib.auth.models import User
from django.db.models import Q
from django.conf import settings
import csv # <--- NEW IMPORT
from django.utils.text import slugify # <--- NEW IMPORT
from celery.result import AsyncResult # <-- ADD CELERY IMPORT
from celery.result import AsyncResult 
from .tasks import process_practice_submission, process_assessment_submission # (and other tasks)
from .models import (
    Notice, NoticeReadStatus, Question, Submission, Module,
    Assessment, AssessmentQuestion, AssessmentSubmission,
    AssessmentSession
)
from .forms import ModuleForm, NoticeForm, QuestionForm
from codingapp import models
from codingapp.models import Department, Role, ActionPermission, UserProfile
from codingapp.utils import deny_access_if_not_allowed, get_user_accessible_groups
from codingapp.utils import is_teacher, is_hod, is_admin, has_role
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest
from django.utils import timezone
import difflib  # for plagiarism
from difflib import SequenceMatcher
import re
from django.contrib import messages
from django.shortcuts import redirect
def is_teacher_or_admin(user):
    if not user.is_authenticated:
        return False
    role = user.userprofile.role.name.lower()
    return role in ["teacher", "admin"]




# Piston (code execution) API endpoint
#PISTON_API_URL = "https://emkc.org/api/v2/piston/execute"
#PISTON_API_URL = "http://localhost:2000/api/v2/execute"
PISTON_API_URL = settings.PISTON_API_URL
SUPPORTED_LANGUAGES = ["python", "c", "cpp", "java", "javascript"]

logger = logging.getLogger(__name__)



from django.contrib.auth.views import LoginView

def is_request_mobile(request):
    """
    Simple server-side mobile detection using User-Agent.
    This isn't perfect but blocks obvious mobile UAs.
    """
    ua = request.META.get('HTTP_USER_AGENT', '') or ''
    ua = ua.lower()
    # basic mobile indicators
    mobile_regex = re.compile(r"android|iphone|ipad|ipod|windows phone|mobile")
    return bool(mobile_regex.search(ua))


class CustomLoginView(LoginView):
    def form_valid(self, form):
        response = super().form_valid(form)
        self.request.session['force_splash'] = True
        return response


def is_admin(user):
    return user.is_staff

import requests
import logging

# Get an instance of a logger
logger = logging.getLogger(__name__)

# def execute_code(code, language, test_cases):
#     """
#     Helper to run code via Piston and collect results, with improved error handling.
#     """
#     results = []
#     error_output = None

#     for test in test_cases or []:
#         payload = {
#             "language": language,
#             "version": "*",
#             "files": [{"name": "solution", "content": code}],
#             "stdin": test.get("input", "")
#         }

#         try:
#             resp = requests.post(PISTON_API_URL, json=payload, timeout=10)
#             resp.raise_for_status()  # Raise an exception for bad status codes (4xx or 5xx)

#             data = resp.json()
#             stdout = data.get("run", {}).get("stdout", "").strip()
#             stderr = data.get("run", {}).get("stderr", "").strip()

#             actual_lines = [line.strip() for line in stdout.splitlines() if line.strip()]
#             expected_lines = test.get("expected_output", [])

#             if stderr:
#                 status = "Error"
#                 error_message = stderr
#             elif len(actual_lines) != len(expected_lines):
#                 status = "Rejected"
#                 error_message = f"Output mismatch: Expected {len(expected_lines)} lines, but got {len(actual_lines)}."
#             else:
#                 mismatches = [f"Line {i+1}: Expected '{expected}', got '{actual}'"
#                               for i, (actual, expected) in enumerate(zip(actual_lines, expected_lines))
#                               if actual != expected]
#                 if mismatches:
#                     status = "Rejected"
#                     error_message = "; ".join(mismatches)
#                 else:
#                     status = "Accepted"
#                     error_message = ""

#             results.append({
#                 "input": test.get("input", ""),
#                 "expected_output": expected_lines,
#                 "actual_output": actual_lines,
#                 "status": status,
#                 "error_message": error_message
#             })

#         except requests.exceptions.Timeout:
#             logger.error("Piston API request timed out.", exc_info=True)
#             msg = "The code execution timed out. Please check for infinite loops or inefficient code."
#             results.append({
#                 "input": test.get("input", ""),
#                 "expected_output": test.get("expected_output", []),
#                 "actual_output": [],
#                 "status": "Error",
#                 "error_message": msg
#             })
#             break

#         except requests.exceptions.HTTPError as http_err:
#             logger.error(f"Piston API returned an HTTP error: {http_err}", exc_info=True)
#             msg = f"An error occurred with the code execution engine (HTTP {http_err.response.status_code})."
#             results.append({
#                 "input": test.get("input", ""),
#                 "expected_output": test.get("expected_output", []),
#                 "actual_output": [],
#                 "status": "Error",
#                 "error_message": msg
#             })
#             break

#         except requests.exceptions.RequestException as e:
#             logger.error(f"Piston API request failed: {e}", exc_info=True)
#             msg = "Could not connect to the code execution engine. Please try again later."
#             results.append({
#                 "input": test.get("input", ""),
#                 "expected_output": test.get("expected_output", []),
#                 "actual_output": [],
#                 "status": "Error",
#                 "error_message": msg
#             })
#             break

#     return results, error_output

def home(request):
    # If the user is authenticated, redirect them to the dashboard.
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    # If not authenticated, render the new welcome page.
    return render(request, 'codingapp/welcome.html') # <--- NEW TARGET TEMPLATE

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from .models import Submission, UserProfile  # Adjust import path as needed

@login_required
def user_dashboard(request):
    # Get the user's submissions
    subs = Submission.objects.filter(user=request.user).order_by('-submitted_at')
    accepted = subs.filter(status="Accepted").count()
    rejected = subs.filter(status="Rejected").count()
    
    # Get the user's profile
    # If you are sure every user has a UserProfile, you can use request.user.userprofile
    # Otherwise, use get_object_or_404 or handle DoesNotExist
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        profile = None  # Or handle as you wish

    return render(request, 'codingapp/dashboard.html', {
        'user_submissions': subs,
        'accepted_count': accepted,
        'rejected_count': rejected,
        'profile': profile,
        'user': request.user,  # Optional, for convenience in template
    })

from django.shortcuts import render, redirect
from django.contrib.auth import login
from .forms import RegistrationForm
from .models import UserProfile

from django.contrib.auth.models import User
from codingapp.forms import StudentRegistrationForm
from codingapp.models import UserProfile

from codingapp.models import Role, UserProfile
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from codingapp.models import UserProfile, Role

from django.shortcuts import render, redirect
from .forms import StudentRegistrationForm
from .utils import generate_otp, send_otp_email
from .models import EmailOTP
from django.core.mail import send_mail

import random
from django.core.mail import send_mail

# codingapp/views.py

import random
from django.core.mail import send_mail
from django.conf import settings
from django.shortcuts import render, redirect
from .forms import RegistrationForm

def register(request):
    if request.method == "POST":
        form = RegistrationForm(request.POST)

        if form.is_valid():
            otp = random.randint(100000, 999999)

            # ⚠️ Store ONLY SERIALIZABLE DATA
            request.session["reg_data"] = {
                "username": form.cleaned_data["username_input"],
                "email": form.cleaned_data["email"],
                "password": form.cleaned_data["password1"],
                "role": form.cleaned_data["role"],
                "department_id": form.cleaned_data["department"].id,
                "group_id": form.cleaned_data["group"].id if form.cleaned_data["group"] else None,
                "otp": str(otp),
            }

            send_mail(
                subject="CodeLoop OTP Verification",
                message=f"Your OTP is {otp}",
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[form.cleaned_data["email"]],
                fail_silently=False,
            )

            return redirect("verify_otp")
    else:
        form = RegistrationForm()

    return render(request, "registration/register.html", {"form": form})




def teacher_register(request):
    if request.method == "POST":
        form = TeacherRegistrationForm(request.POST)

        if form.is_valid():
            email = form.cleaned_data["email"]
            username = form.cleaned_data["username"]

            if User.objects.filter(username=username).exists():
                form.add_error("username", "Username already exists.")
                return render(request, "registration/teacher_register.html", {"form": form})

            user = User.objects.create_user(
                username=username,
                email=email,
                password=form.cleaned_data["password1"]
            )

            teacher_role, _ = Role.objects.get_or_create(name="teacher")

            UserProfile.objects.get_or_create(
                user=user,
                defaults={
                    "role": teacher_role,
                    "department": form.cleaned_data["department"]
                }
            )

            return redirect("login")

    else:
        form = TeacherRegistrationForm()

    return render(request, "registration/teacher_register.html", {"form": form})


from django.contrib.auth.models import User
from .models import Role, UserProfile, Group, Department

from django.contrib.auth.models import User
from codingapp.models import UserProfile, Role, Department, Group

from django.contrib.auth.models import User
from .models import Role, UserProfile, Group, Department

from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import UserProfile, Role, Group, Department

def verify_otp(request):
    if request.method == "POST":
        entered_otp = request.POST.get("otp")
        reg_data = request.session.get("reg_data")

        if not reg_data:
            messages.error(request, "Session expired. Please register again.")
            return redirect("register")

        if entered_otp != reg_data.get("otp"):
            messages.error(request, "Invalid OTP")
            return render(request, "registration/verify_otp.html")

        # ✅ OTP VERIFIED — CREATE ACCOUNT SAFELY
        try:
            with transaction.atomic():

                # 1️⃣ Create User
                user, created = User.objects.get_or_create(
                    username=reg_data["username"],
                    defaults={
                        "email": reg_data["email"],
                    }
                )

                if not created:
                    messages.error(request, "User already exists.")
                    return redirect("register")

                user.set_password(reg_data["password"])
                user.save()

                # 2️⃣ Fetch Role instance
                role = Role.objects.get(name__iexact=reg_data["role"])

                department = Department.objects.get(id=reg_data["department_id"])

                # 3️⃣ Create UserProfile SAFELY
                profile, _ = UserProfile.objects.get_or_create(
                    user=user,
                    defaults={
                        "role": role,
                        "department": department,
                    }
                )

                # 4️⃣ Assign group ONLY for students
                if reg_data.get("group_id"):
                    group = Group.objects.get(id=reg_data["group_id"])
                    group.students.add(user)

                # 5️⃣ Cleanup session
                del request.session["reg_data"]

                messages.success(request, "Account verified and created successfully.")
                return redirect("login")

        except Exception as e:
            messages.error(request, f"Registration failed: {str(e)}")
            return redirect("register")

    return render(request, "registration/verify_otp.html")


import random
from django.core.mail import send_mail
from django.contrib import messages
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from .forms import ForgotPasswordForm, ResetPasswordForm

def forgot_password(request):
    if request.method == "POST":
        form = ForgotPasswordForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]

            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                messages.error(request, "No account found with this email.")
                return redirect("forgot_password")

            otp = str(random.randint(100000, 999999))

            request.session["fp_otp"] = otp
            request.session["fp_user_id"] = user.id

            send_mail(
                subject="Password Reset OTP - CodeLoop",
                message=f"Your OTP for password reset is: {otp}",
                from_email=None,
                recipient_list=[email],
            )

            messages.success(request, "OTP sent to your email.")
            return redirect("reset_password")

    else:
        form = ForgotPasswordForm()

    return render(request, "registration/forgot_password.html", {"form": form})

def reset_password(request):
    if request.method == "POST":
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            otp_entered = form.cleaned_data["otp"]
            session_otp = request.session.get("fp_otp")
            user_id = request.session.get("fp_user_id")

            if not session_otp or not user_id:
                messages.error(request, "Session expired.")
                return redirect("forgot_password")

            if otp_entered != session_otp:
                messages.error(request, "Invalid OTP.")
                return redirect("reset_password")

            user = User.objects.get(id=user_id)
            user.set_password(form.cleaned_data["new_password"])
            user.save()

            # Cleanup
            del request.session["fp_otp"]
            del request.session["fp_user_id"]

            messages.success(request, "Password reset successful. Please login.")
            return redirect("login")

    else:
        form = ResetPasswordForm()

    return render(request, "registration/reset_password.html", {"form": form})



from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, get_object_or_404, redirect
from .models import Module, Submission, Question
from .forms import ModuleForm, QuestionForm

@login_required
def module_list(request):
    accessible_groups = get_user_accessible_groups(request.user)
    mods = Module.objects.filter(groups__in=accessible_groups).distinct()
    return render(request, 'codingapp/module_list.html', {'modules': mods})


from .models import ModuleCompletion  # make sure this is imported

@login_required
def module_detail(request, module_id):
    module = get_object_or_404(Module, id=module_id)
    denied = deny_access_if_not_allowed(request, module)
    if denied: 
        return denied
    if not request.user.is_staff:
        user_groups = request.user.custom_groups.all()
        if not module.groups.filter(id__in=user_groups.values_list('id', flat=True)).exists():
            return render(request, "codingapp/permission_denied.html", status=403)

    questions = module.questions.all()
    total_count = questions.count()

    completed_count = Submission.objects.filter(
        user=request.user,
        question__in=questions,
        status="Accepted"
    ).values("question").distinct().count()

    # ✅ Check if already marked as completed
    is_completed = ModuleCompletion.objects.filter(
        user=request.user,
        module=module
    ).exists()

    return render(request, "codingapp/module_detail.html", {
        "module": module,
        "total_count": total_count,
        "completed_count": completed_count,
        "is_completed": is_completed,  # ✅ Add to context
    })


@user_passes_test(is_teacher, login_url='/dashboard/')
def add_module(request):
    if request.method == 'POST':
        form = ModuleForm(request.POST)
        if form.is_valid():
            module = form.save()  # save module + groups correctly
            messages.success(request, "Module created successfully!")
            return redirect("module_list")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ModuleForm()

    return render(request, "codingapp/module_form.html", {
        "form": form,
        "action": "Add",
    })


@user_passes_test(is_teacher, login_url='/dashboard/')
def edit_module(request, module_id):
    print("### USING edit_module VIEW ###")

    module = get_object_or_404(Module, id=module_id)

    if request.method == 'POST':
        print("\n===== DEBUG: EDIT MODULE POST RECEIVED =====")
        print("POST DATA:", request.POST)

        form = ModuleForm(request.POST, instance=module)

        print("Form is_valid():", form.is_valid())
        print("Form.errors:", form.errors)
        print("Non field errors:", form.non_field_errors())

        if form.is_valid():
            saved = form.save()
            print("MODULE SAVED:", saved)
            messages.success(request, "Module updated successfully!")
            return redirect("module_list")
        else:
            print("DEBUG: Form did NOT validate.")

    else:
        form = ModuleForm(instance=module)

    return render(request, "codingapp/module_form.html", {
        "form": form,
        "action": "Edit",
    })



@user_passes_test(is_teacher, login_url='/dashboard/')
def delete_module(request, module_id):
    mod = get_object_or_404(Module, id=module_id)
    if request.method == "POST":
        mod.delete()
        return redirect("module_list")
    return render(request, "codingapp/module_confirm_delete.html", {"module": mod})

@login_required
def question_list(request):
    accessible_groups = get_user_accessible_groups(request.user)
    modules = Module.objects.filter(groups__in=accessible_groups).distinct()
    qs = Question.objects.filter(module__in=modules, question_type="coding").distinct()

    return render(request, 'codingapp/question_list.html', {'questions': qs})


@login_required
def question_detail(request, pk):
    from celery.result import AsyncResult
    import json
    from django.conf import settings

    q = get_object_or_404(Question, pk=pk)
    denied = deny_access_if_not_allowed(request, q.module)
    if denied:
        return denied

    # --- Access control ---
    if not request.user.is_staff:
        user_groups = request.user.custom_groups.all()
        if not q.module or not q.module.groups.filter(id__in=user_groups.values_list('id', flat=True)).exists():
            return render(request, "codingapp/permission_denied.html", status=403)

    # -------------------------
    # Initialize variables
    # -------------------------
    results = None
    error = None

    # --- Determine currently selected language ---
    if request.method == "POST":
        selected_lang = request.POST.get("language", "python")
    else:
        selected_lang = request.session.get(f'last_language_{pk}', request.session.get(f'language_{pk}', 'python'))

    # Session key patterns (store separately for each language)
    session_code_key = f'code_{pk}_{selected_lang}'
    session_task_key = f'submission_task_id_{pk}_{selected_lang}'

    # Load existing task/code from session (if any)
    task_id = request.session.get(session_task_key)
    code = request.session.get(session_code_key, '')

    # -------------------------
    # 1️⃣ Handle Submission (POST)
    # -------------------------
    if request.method == "POST":
        code = request.POST.get("code", "").strip()
        lang = request.POST.get("language", "python")

        if not code:
            messages.error(request, "Code cannot be empty.")
        else:
            # ⭐ Run async submission
            task = process_practice_submission.delay(request.user.id, q.id, code, lang)

            # Save per-language info in session
            request.session[f'submission_task_id_{pk}_{lang}'] = task.id
            request.session[f'code_{pk}_{lang}'] = code
            request.session[f'last_language_{pk}'] = lang
            request.session[f'language_{pk}'] = lang
            request.session.modified = True

            messages.info(request, "Your code is being processed in the background. Please wait or check back shortly.")
            return redirect('question_detail', pk=pk)

    # -------------------------
    # 2️⃣ Handle async task status (GET)
    # -------------------------
    if task_id:
        task_result = AsyncResult(task_id)

        if task_result.ready():
            # Remove task marker once finished
            request.session.pop(session_task_key, None)
            request.session.modified = True

            # Fetch last DB submission for this language
            latest_submission = Submission.objects.filter(
                user=request.user,
                question=q,
                language=selected_lang
            ).order_by('-submitted_at').first()

            if latest_submission:
                try:
                    results = json.loads(latest_submission.output) if latest_submission.output else None
                except json.JSONDecodeError:
                    results = None

                if latest_submission.status == "Accepted":
                    messages.success(request, "✅ Code Accepted! All test cases passed.")
                elif latest_submission.status == "Error":
                    messages.error(request, f"❌ Code Execution Error: {latest_submission.error}")
                    error = latest_submission.error
                else:
                    messages.warning(request, "⚠️ Some test cases failed. Try again.")
            task_id = None

        else:
            messages.info(request, f"Submission is processing... Status: {task_result.status}")

    # -------------------------
    # 3️⃣ Load code if not in session
    # -------------------------
    if not code:
        last_submission = Submission.objects.filter(
            user=request.user,
            question=q,
            language=selected_lang
        ).order_by('-submitted_at').first()

        if last_submission:
            code = last_submission.code
            if not results and last_submission.output:
                try:
                    results = json.loads(last_submission.output)
                except json.JSONDecodeError:
                    pass

    # -------------------------
    # 4️⃣ Build user_submissions map (for all languages)
    # -------------------------
    # Collect DB submissions
    db_codes = {
        s.language: s.code
        for s in Submission.objects.filter(user=request.user, question=q)
    }

    # Merge with session codes (to include unsaved async submissions)
    session_codes = {}
    for key, val in request.session.items():
        if key.startswith(f'code_{pk}_'):
            lang_name = key.split(f'code_{pk}_', 1)[1]
            session_codes[lang_name] = val

    user_submissions = {**db_codes, **session_codes}

    # -------------------------
    # 5️⃣ Prepare context and render
    # -------------------------
    context = {
        "question": q,
        "code": code,
        "selected_language": selected_lang,
        "results": results,
        "error": error,
        "task_id": task_id,
        "supported_languages": settings.SUPPORTED_LANGUAGES,
        "user_submissions": user_submissions,
    }

    return render(request, "codingapp/question_detail.html", context)


from django.views.decorators.http import require_POST
from .models import Module, ModuleCompletion, Submission
from django.contrib import messages

@login_required
@require_POST
def mark_module_completed(request, module_id):
    module = get_object_or_404(Module, id=module_id)

    if not request.user.is_staff:
        user_groups = request.user.custom_groups.all()
        if not module.groups.filter(id__in=user_groups.values_list('id', flat=True)).exists():
            return render(request, "codingapp/permission_denied.html", status=403)

    questions = module.questions.all()
    total_questions = questions.count()

    completed_questions = Submission.objects.filter(
        user=request.user,
        question__in=questions,
        status='Accepted'
    ).values('question').distinct().count()

    if total_questions > 0 and completed_questions == total_questions:
        # Only create if not already marked
        ModuleCompletion.objects.get_or_create(user=request.user, module=module)
        messages.success(request, "Module marked as completed!")
    else:
        messages.error(request, "You must complete all questions (Accepted) before marking as completed.")

    return redirect('module_detail', module_id=module.id)



from .forms import QuestionForm, TestCaseForm
from django.forms import formset_factory
from django.contrib.admin.views.decorators import staff_member_required
@user_passes_test(is_teacher, login_url='/dashboard/')
def add_question_to_module(request, module_id):
    mod = get_object_or_404(Module, id=module_id)
    TestCaseFormSet = formset_factory(TestCaseForm, extra=1, can_delete=True)
    if request.method == "POST":
        form = QuestionForm(request.POST)
        formset = TestCaseFormSet(request.POST, prefix='testcases')
        if form.is_valid() and (form.cleaned_data['question_type'] != 'coding' or formset.is_valid()):
            q = form.save(commit=False)
            q.module = mod
            if form.cleaned_data['question_type'] == 'coding':
                test_cases = []
                for f in formset.cleaned_data:
                    if f and not f.get('DELETE', False):
                        test_cases.append({
                            'input': f['input'],
                            'expected_output': [line.strip() for line in f['expected_output'].splitlines() if line.strip()],
                        })
                if not test_cases:
                    # Validation: At least one test case required
                    formset.non_form_errors = ["At least one test case is required for coding questions."]
                    return render(request, 'codingapp/add_question.html', {
                        'form': form, 'formset': formset, 'module': mod,
                    })
                q.test_cases = test_cases
            else:
                q.test_cases = []
            q.save()
            return redirect('module_detail', module_id=mod.id)
    else:
        form = QuestionForm()
        formset = TestCaseFormSet(prefix='testcases')
    return render(request, 'codingapp/add_question.html', {
        'form': form,
        'formset': formset,
        'module': mod,
    })


def leaderboard(request):
    top = (User.objects
           .annotate(accepted_count=Count('submission', filter=Q(submission__status="Accepted")))
           .filter(accepted_count__gt=0)
           .order_by('-accepted_count'))
    return render(request, "codingapp/leaderboard.html", {"top_users": top})

from django.utils import timezone

@login_required
def assessment_list(request):
    now = timezone.now()
    accessible_groups = get_user_accessible_groups(request.user)
    asses = Assessment.objects.filter(end_time__gte=now, groups__in=accessible_groups).distinct()

    return render(request, 'codingapp/assessment_list.html', {"assessments": asses})

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from .models import Assessment, AssessmentSession, AssessmentQuestion

@login_required
def assessment_detail(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    # Basic permission and active checks
    if not request.user.is_staff:
        user_groups = request.user.custom_groups.all()
        if not assessment.groups.filter(id__in=user_groups.values_list('id', flat=True)).exists():
            return render(request, "codingapp/permission_denied.html", status=403)
    if not assessment.is_active():
        messages.warning(request, "This assessment is not active.")
        return redirect("assessment_list")
        # block mobile devices from starting the assessment (server-side)
    if is_request_mobile(request):
        messages.error(request, "Assessments cannot be taken on mobile devices. Please use a desktop or laptop.")
        return redirect('assessment_list')   # or 'assessment_detail' based on UX

    # Get or create the session
    session, created = AssessmentSession.objects.get_or_create(
        user=request.user,
        assessment=assessment,
        defaults={"start_time": timezone.now()}
    )

    # --- FIX: Check if the user has already finished the assessment ---
    if session.end_time:
        messages.info(request, "You have already completed this assessment.")
        return redirect('assessment_result', assessment_id=assessment.id)

    # Redirect to the quiz if it exists and hasn't been submitted
    if assessment.quiz and not session.quiz_submitted:
        return redirect('assessment_quiz', assessment_id=assessment.id)

    # If no quiz or quiz is done, redirect to the first coding question
    first_question = AssessmentQuestion.objects.filter(assessment=assessment).order_by('order').first()
    if first_question:
        return redirect('submit_assessment_code', assessment_id=assessment.id, question_id=first_question.question.id)

    # Fallback if there are no coding questions
    messages.error(request, "This assessment has no coding questions.")
    return redirect('assessment_list')


@login_required
def assessment_quiz(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    session = get_object_or_404(AssessmentSession, user=request.user, assessment=assessment)
    quiz = assessment.quiz

    if session.quiz_submitted:
        # If the quiz is already done, redirect to the main assessment handler
        return redirect('assessment_detail', assessment_id=assessment.id)

    # --- FIX STARTS HERE ---
    # 1. Calculate the assessment's deadline to pass to the template's timer
    deadline = session.start_time + timezone.timedelta(minutes=assessment.duration_minutes)
    
    # 2. Prepare the list of coding questions for the navigation bar
    # This ensures the nav bar is visible even on the quiz page for a consistent UI
    all_coding_questions = AssessmentQuestion.objects.filter(assessment=assessment).select_related('question').order_by('order')
    nav_questions_data = []
    for aq in all_coding_questions:
        # At the quiz stage, no coding questions have been attempted or solved yet
        nav_questions_data.append({
            'question': aq.question,
            'is_solved': False,
            'is_attempted': False,
        })
    # --- FIX ENDS HERE ---

    # Handle the quiz logic
    questions = list(quiz.questions.all())
    import random
    random.shuffle(questions)

    if request.method == 'POST':
        submission = QuizSubmission.objects.create(user=request.user, quiz=quiz)
        score = 0
        for q in questions:
            selected = request.POST.get(f'question_{q.id}')
            if selected:
                QuizAnswer.objects.create(submission=submission, question=q, selected_option=selected)
                if selected == q.correct_answer:
                    score += 1
        submission.score = score
        submission.save()
        
        session.quiz_submitted = True
        session.save()
        
        messages.success(request, f"Quiz submitted! Your score: {score}/{len(questions)}. The coding section is now unlocked.")
        return redirect('assessment_detail', assessment_id=assessment.id)

    # 3. Add the new data to the context dictionary
    context = {
        'assessment': assessment,
        'quiz': quiz,
        'questions': questions,
        'end_time': deadline.isoformat(),      # This fixes the timer
        'all_questions': nav_questions_data,  # This provides data for the nav bar
        'focus_mode': True,  # 👈 ADD THIS LINE
    }
    
    return render(request, 'codingapp/assessment_quiz.html', context)

@login_required
def assessment_leaderboard_list(request):
    """
    Displays a list of all assessments for which a leaderboard can be viewed.
    """
    if request.user.is_staff:
        # Staff can see leaderboards for all assessments
        assessments = Assessment.objects.all().order_by('-end_time')
    else:
        # Students can see leaderboards for assessments in their groups
        user_groups = request.user.custom_groups.all()
        assessments = Assessment.objects.filter(groups__in=user_groups).distinct().order_by('-end_time')
    
    context = {
        'assessments': assessments
    }
    return render(request, 'codingapp/assessment_leaderboard_list.html', context)

# In codingapp/views.py

# at top of file (if not already)
from django.db.models import Sum, Max, F
from django.contrib.auth.models import User
from .models import Assessment, AssessmentSubmission, AssessmentSession, QuizSubmission

from django.db.models import Sum, Max

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.db.models import Sum, Max
from django.contrib.auth import get_user_model

from .models import (
    Assessment, AssessmentQuestion, AssessmentSubmission, AssessmentSession,
    QuizSubmission
)
from .utils import penalty_factor_from_plagiarism  # make sure this exists

User = get_user_model()

@login_required
def assessment_leaderboard(request, assessment_id):
    from django.db.models import Sum, Max
    import json
    from django.shortcuts import get_object_or_404, render
    from django.contrib.auth import get_user_model

    # local model imports to avoid top-level import issues
    from codingapp.models import (
        Assessment, AssessmentQuestion, AssessmentSubmission,
        AssessmentSession, QuizSubmission
    )

    User = get_user_model()

    assessment = get_object_or_404(Assessment, id=assessment_id)

    # Permission Check
    if not request.user.is_staff:
        user_groups = request.user.custom_groups.all()
        if not assessment.groups.filter(id__in=user_groups.values_list('id', flat=True)).exists():
            return render(request, "codingapp/permission_denied.html", status=403)

    # =========================================================
    # 1. CHANGED: Participants (Filtered strictly by Assigned Group)
    # =========================================================
    # This ensures we ONLY get students (not staff) who belong to 
    # the groups specifically assigned to this assessment.
    participants = User.objects.filter(
        custom_groups__in=assessment.groups.all(),
        is_staff=False
    ).distinct()
    # =========================================================

    # 2. Questions order
    assessment_questions = (
        AssessmentQuestion.objects
        .filter(assessment=assessment)
        .select_related('question')
        .order_by('order')
    )
    questions_list = [aq.question for aq in assessment_questions]

    # 3. All submissions for this assessment (select related for fewer queries)
    all_submissions = AssessmentSubmission.objects.filter(assessment=assessment).select_related('question', 'user')

    # helper: normalize status strings that indicate a passed test
    def _is_accepted_status(s):
        if not s:
            return False
        s = str(s).strip().lower()
        return s in ("accepted", "ok", "passed", "success", "true")

    # Precompute question test counts where available
    question_testcounts = {}
    for q in questions_list:
        try:
            question_testcounts[q.id] = len(q.test_cases or [])
        except Exception:
            question_testcounts[q.id] = None

    # Build maps of latest submission per user-question (we'll iterate in time order so later submissions overwrite)
    user_question_scores = {}  # penalized scores stored in DB (sub.score)
    user_question_raw = {}     # raw marks inferred or stored (0..5)
    user_question_extra = {}   # for ai/token/struct similarity per submission (latest)

    # iterate submissions in time order (older -> newer) so later ones overwrite
    for sub in all_submissions.order_by('submitted_at'):
        uid = sub.user_id
        qid = sub.question_id

        # penalized score as stored
        try:
            penalized_val = float(sub.score or 0)
        except Exception:
            penalized_val = 0.0
        user_question_scores.setdefault(uid, {})[qid] = penalized_val

        # prefer explicit raw_score if present
        raw_val = getattr(sub, "raw_score", None)
        if raw_val is not None:
            try:
                raw_numeric = float(raw_val or 0.0)
            except Exception:
                raw_numeric = 0.0
            user_question_raw.setdefault(uid, {})[qid] = raw_numeric
        else:
            # try to infer from output JSON/list/dict
            inferred_raw = 0.0
            outputs = []
            try:
                out_field = sub.output
                if isinstance(out_field, str):
                    # try safe json first
                    try:
                        parsed = json.loads(out_field)
                        if isinstance(parsed, list):
                            outputs = parsed
                        elif isinstance(parsed, dict):
                            outputs = [parsed]
                        else:
                            outputs = []
                    except Exception:
                        # not valid json -> keep empty list (fallback to score)
                        outputs = []
                elif isinstance(out_field, list):
                    outputs = out_field
                elif isinstance(out_field, dict):
                    outputs = [out_field]
                else:
                    outputs = []
            except Exception:
                outputs = []

            # Count passed testcases
            passed = 0
            for r in outputs:
                try:
                    # Accept either 'status' or 'result' etc.
                    if isinstance(r, dict):
                        st = r.get("status") or r.get("result") or r.get("outcome")
                    else:
                        st = None
                    if _is_accepted_status(st):
                        passed += 1
                except Exception:
                    continue

            total_tests = question_testcounts.get(qid)
            if total_tests is None or total_tests == 0:
                total_tests = len(outputs) or 1

            # policy: full 5 only if all passed; else 0 (you can switch to proportional easily)
            if total_tests > 0 and passed == total_tests:
                inferred_raw = 5.0
            else:
                inferred_raw = 0.0

            user_question_raw.setdefault(uid, {})[qid] = float(inferred_raw)

        # store extra similarity metrics if present on submission (latest overwrite)
        user_question_extra.setdefault(uid, {})[qid] = {
            'token_similarity': float(getattr(sub, 'token_similarity', 0.0) or 0.0),
            'structural_similarity': float(getattr(sub, 'structural_similarity', 0.0) or 0.0),
            'ai_generated_prob': float(getattr(sub, 'ai_generated_prob', 0.0) or 0.0),
            'plagiarism_percent': float(getattr(sub, 'plagiarism_percent', 0.0) or 0.0),
        }

    # 4. Total (penalized) coding scores aggregated
    coding_scores_agg = (
        all_submissions
        .values('user__id')
        .annotate(total_coding_score=Sum('score'))
    )
    coding_scores_dict = { item['user__id']: float(item['total_coding_score'] or 0.0) for item in coding_scores_agg }

    # 5. Max plagiarism percent per user (DB aggregate)
    plag_agg = (
        all_submissions
        .values('user__id')
        .annotate(max_plag=Max('plagiarism_percent'))
    )
    plagiarism_dict = { item['user__id']: float(item['max_plag'] or 0.0) for item in plag_agg }

    # 6. Quiz scores (best per user)
    quiz_scores_dict = {}
    if getattr(assessment, "quiz", None):
        # try to gather quiz submissions; be defensive about model shape
        try:
            quiz_subs = QuizSubmission.objects.filter(quiz=assessment.quiz).values('user__id', 'score')
            for item in quiz_subs:
                uid = item['user__id']
                sc = float(item['score'] or 0.0)
                quiz_scores_dict[uid] = max(quiz_scores_dict.get(uid, 0.0), sc)
        except Exception:
            # if QuizSubmission not present or relation differs, skip quietly
            quiz_scores_dict = {}

    # 7. Sessions & time taken
    sessions = AssessmentSession.objects.filter(
        assessment=assessment,
        user__in=participants,
        end_time__isnull=False,
    ).select_related('user')

    session_times_dict = {}
    for session in sessions:
        if session.start_time and session.end_time:
            duration = session.end_time - session.start_time
            total_seconds = int(duration.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            time_str = f"{hours}h {minutes}m {seconds}s"
            session_times_dict[session.user_id] = {
                'time_str': time_str,
                'sort_key': total_seconds,
                'session': session
            }
        else:
            session_times_dict[session.user_id] = {
                'time_str': "N/A",
                'sort_key': float('inf'),
                'session': session
            }

    # 8. Build leaderboard rows
    leaderboard_data = []
    for user in participants:
        uid = user.id
        question_pairs = []
        # Build per-question pairs
        for q in questions_list:
            penalized_val = float(user_question_scores.get(uid, {}).get(q.id, 0.0) or 0.0)
            raw_val = float(user_question_raw.get(uid, {}).get(q.id, 0.0) or 0.0)
            question_pairs.append({
                'penalized': penalized_val,
                'raw': raw_val,
            })

        total_coding = float(coding_scores_dict.get(uid, 0.0) or 0.0)
        total_coding_raw = float(sum(item['raw'] for item in question_pairs) or 0.0)
        quiz_score = float(quiz_scores_dict.get(uid, 0.0) or 0.0)
        max_plag = round(plagiarism_dict.get(uid, 0.0), 2)
        raw_total = round(quiz_score + total_coding_raw, 2)

        # Compute per-user maxes for similarity metrics across questions (latest per-sub overwrites)
        max_token = 0.0
        max_struct = 0.0
        max_ai = 0.0
        for q in questions_list:
            extra = user_question_extra.get(uid, {}).get(q.id, {})
            max_token = max(max_token, float(extra.get('token_similarity', 0.0) or 0.0))
            max_struct = max(max_struct, float(extra.get('structural_similarity', 0.0) or 0.0))
            max_ai = max(max_ai, float(extra.get('ai_generated_prob', 0.0) or 0.0))

        # Session-level penalty fields if stored
        session = session_times_dict.get(uid, {}).get('session')
        if session is not None and hasattr(session, "penalized_total") and session.penalized_total is not None:
            try:
                penalized_total = float(session.penalized_total)
            except Exception:
                penalized_total = raw_total
            penalty_factor = float(getattr(session, "penalty_factor", 1.0) or 1.0)
            penalty_applied = bool(getattr(session, "penalty_applied", penalty_factor < 1.0))
        else:
            # compute on-the-fly from max_plag
            from codingapp.utils import penalty_factor_from_plagiarism
            try:
                penalty_factor = penalty_factor_from_plagiarism(max_plag)
            except Exception:
                if max_plag <= 35.0:
                    penalty_factor = 1.0
                elif max_plag >= 85.0:
                    penalty_factor = 0.0
                else:
                    penalty_factor = round((85.0 - max_plag) / (85.0 - 35.0), 4)
            penalized_total = round(raw_total * penalty_factor, 2)
            penalty_applied = (penalty_factor < 1.0)

        total_score_display = penalized_total

        leaderboard_data.append({
            'user_id': uid,
            'username': user.get_full_name() or user.username,
            'question_pairs': question_pairs,
            'total_coding_score': total_coding,
            'total_coding_raw': total_coding_raw,
            'quiz_score': quiz_score,
            'raw_total': raw_total,
            'penalized_total': penalized_total,
            'penalty_factor': penalty_factor,
            'penalty_applied': penalty_applied,
            'total_score': total_score_display,
            'time_taken_str': session_times_dict.get(uid, {'time_str': "N/A"})['time_str'],
            'time_taken_seconds': session_times_dict.get(uid, {'sort_key': float('inf')})['sort_key'],
            'max_plagiarism': max_plag,
            # Convert similarity decimals into percentages for display (keep a few decimals)
            'token_similarity': round(max_token * 100.0, 4),
            'structural_similarity': round(max_struct * 100.0, 4),
            'ai_generated_prob': round(max_ai * 100.0, 4),
        })

    # 9. Sort leaderboard
    leaderboard_data.sort(key=lambda row: (-row['total_score'], row['time_taken_seconds']))

    context = {
        'assessment': assessment,
        'questions': questions_list,
        'leaderboard': leaderboard_data,
    }
    return render(request, 'codingapp/assessment_leaderboard.html', context) 



from django.contrib import messages

from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponseBadRequest
from django.utils import timezone

@csrf_exempt
@require_POST
@login_required
def assessment_heartbeat(request, assessment_id):
    try:
        data = json.loads(request.body.decode('utf-8') or "{}")
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    in_fullscreen = data.get("in_fullscreen")
    if in_fullscreen is None:
        return HttpResponseBadRequest("Missing 'in_fullscreen' field")

    try:
        session = AssessmentSession.objects.get(user=request.user, assessment_id=assessment_id)
    except AssessmentSession.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Session not found"}, status=404)

    session.last_heartbeat = timezone.now()

    # Only increment warnings if client reports not in fullscreen
    if not bool(in_fullscreen):
        # increment (ensure warnings_count default is numeric)
        session.warnings_count = (session.warnings_count or 0) + 1
        if session.warnings_count > 1:
            session.flagged = True
            if not session.end_time:
                session.end_time = timezone.now()
    session.save()

    return JsonResponse({
        "ok": True,
        "warnings_count": session.warnings_count,
        "flagged": session.flagged
    })


@login_required
def submit_assessment_code(request, assessment_id, question_id):
    # kept local imports (as in your style)
    from celery.result import AsyncResult
    import json
    from django.conf import settings
    from django.db.models import Max
    from django.http import JsonResponse  # already imported above for heartbeat, but safe

    assessment = get_object_or_404(Assessment, id=assessment_id)
    denied = deny_access_if_not_allowed(request, assessment)
    if denied:
        return denied

    current_question_obj = get_object_or_404(Question, id=question_id)

    # ✅ Permission check (unchanged)
    if not request.user.is_staff:
        user_groups = request.user.custom_groups.all()
        if not assessment.groups.filter(id__in=user_groups.values_list('id', flat=True)).exists():
            return render(request, "codingapp/permission_denied.html", status=403)

    # Server-side mobile prevention (defense-in-depth)
    if is_request_mobile(request):
        # For AJAX, return JSON; for normal, redirect
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse(
                {"ok": False, "error": "Coding assessments are not allowed on mobile devices. Use a desktop/laptop."},
                status=400,
            )
        messages.error(request, "Coding assessments are not allowed on mobile devices. Please use a desktop or laptop.")
        return redirect('assessment_detail', assessment_id=assessment.id)

    session = get_object_or_404(AssessmentSession, user=request.user, assessment=assessment)

    # Immediately block interaction if the session was flagged by server-side heartbeat
    if session.flagged:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse(
                {"ok": False, "error": "Your assessment has been ended due to focus violations."},
                status=400,
            )
        messages.error(request, "Your assessment has been ended due to focus violations.")
        return redirect('assessment_result', assessment_id=assessment.id)

    if assessment.quiz and not session.quiz_submitted:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse(
                {"ok": False, "error": "You must complete the quiz section before accessing coding questions."},
                status=400,
            )
        messages.error(request, "You must complete the quiz section before accessing coding questions.")
        return redirect('assessment_detail', assessment_id=assessment.id)

    # Timing logic
    deadline = session.start_time + timezone.timedelta(minutes=assessment.duration_minutes)
    read_only = timezone.now() > deadline

    # ===============================
    # Multi-language setup
    # ===============================
    if request.method == "POST":
        selected_lang = request.POST.get("language", "python")
    else:
        selected_lang = request.GET.get(
            "language",
            request.session.get(
                f'assessment_lang_{assessment_id}_{question_id}',
                'python'
            )
        )

    # Per-language session keys (used for restoring code per language)
    session_code_key = f'assessment_code_{assessment_id}_{question_id}_{selected_lang}'

    # For old session-based polling we used a task key; now polling is pure AJAX
    code = request.session.get(session_code_key, '')
    submission_results = None
    latest_submission = None

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

    # ===============================
    # 1️⃣ AJAX POST – create Celery task
    # ===============================
    if is_ajax and request.method == "POST":
        if read_only:
            return JsonResponse(
                {"ok": False, "error": "Assessment time is over; you cannot submit more code."},
                status=400,
            )

        # We send FormData from JS, so csrf + fields are in request.POST
        code = (request.POST.get("code") or "").strip()
        lang = request.POST.get("language", "python")

        if not code:
            return JsonResponse({"ok": False, "error": "Code cannot be empty."}, status=400)

        # Fire async task
        task = process_assessment_submission.delay(
            request.user.id, assessment.id, current_question_obj.id, code, lang
        )

        # Remember code + language in session for later page loads
        request.session[f'assessment_code_{assessment_id}_{question_id}_{lang}'] = code
        request.session[f'assessment_lang_{assessment_id}_{question_id}'] = lang
        request.session.modified = True

        return JsonResponse(
            {
                "ok": True,
                "task_id": task.id,
                "message": "Submission is being processed in the background.",
            }
        )

    # ===============================
    # 2️⃣ AJAX GET – poll Celery for this task_id
    # ===============================
    if is_ajax and request.method == "GET":
        task_id = request.GET.get("task_id")
        if not task_id:
            return JsonResponse({"ok": False, "error": "Missing task_id"}, status=400)

        result = AsyncResult(task_id)

        # Not ready yet – tell JS to keep polling
        if not result.ready():
            return JsonResponse(
                {
                    "ok": True,
                    "completed": False,
                    "state": result.state,
                }
            )

        # Ready – unpack Celery result dict
        data = result.result or {}
        # In your task you return: final_status, results, error, plagiarism_percent
        final_status = data.get("final_status") or data.get("status") or "Unknown"
        results = data.get("results", [])
        error_msg = data.get("error", "")
        plagiarism_percent = data.get("plagiarism_percent", 0.0)

        return JsonResponse(
            {
                "ok": True,
                "completed": True,
                "state": result.state,
                "final_status": final_status,
                "results": results,
                "error": error_msg,
                "plagiarism_percent": plagiarism_percent,
            }
        )

    # ===============================
    # 3️⃣ Normal GET – show last saved result (from DB)
    # ===============================
    if not code:
        latest_submission = AssessmentSubmission.objects.filter(
            assessment=assessment,
            question=current_question_obj,
            user=request.user,
            language=selected_lang
        ).order_by('-submitted_at').first()

        if latest_submission:
            code = latest_submission.code
            if latest_submission.output and not submission_results:
                try:
                    submission_results = json.loads(latest_submission.output)
                except json.JSONDecodeError:
                    submission_results = None

    # ===============================
    # 4️⃣ Determine read-only state (deadline or full score)
    # ===============================
    best_score_agg = AssessmentSubmission.objects.filter(
        assessment=assessment,
        question=current_question_obj,
        user=request.user
    ).aggregate(best=Max('score'))
    best_score = (best_score_agg.get('best') or 0)

    is_fully_solved = (best_score >= 5)
    final_read_only = read_only or is_fully_solved

    # ===============================
    # 5️⃣ Build user_submissions map (DB + session) for language switching
    # ===============================
    db_codes = {
        s.language: s.code
        for s in AssessmentSubmission.objects.filter(
            user=request.user,
            assessment=assessment,
            question=current_question_obj
        )
    }

    session_codes = {}
    prefix = f'assessment_code_{assessment_id}_{question_id}_'
    for key, val in request.session.items():
        if key.startswith(prefix):
            lang_name = key[len(prefix):]
            session_codes[lang_name] = val

    user_submissions = {**db_codes, **session_codes}

    # ===============================
    # 6️⃣ Navigation Bar
    # ===============================
    all_qs = AssessmentQuestion.objects.filter(assessment=assessment).select_related('question').order_by('order')
    nav_questions = []
    for aq in all_qs:
        sub = AssessmentSubmission.objects.filter(
            assessment=assessment, question=aq.question, user=request.user
        ).first()
        is_solved = sub and sub.score >= 5
        is_attempted = sub is not None
        nav_questions.append({
            'question': aq.question,
            'is_solved': is_solved,
            'is_attempted': is_attempted,
        })

    context = {
        "assessment": assessment,
        "question": current_question_obj,
        "code": code,
        "selected_language": selected_lang,
        "supported_languages": settings.SUPPORTED_LANGUAGES,
        "read_only": final_read_only,
        "end_time": deadline.isoformat(),
        "all_questions": nav_questions,
        "results": submission_results,      # last saved results for initial render
        "focus_mode": True,
        "user_submissions": user_submissions,
    }

    return render(request, "codingapp/submit_assessment_code.html", context)

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import user_passes_test
from .models import Module
from .forms import ModuleForm


@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_module_list(request):
    modules = Module.objects.all()
    return render(request, 'codingapp/teacher_module_list.html', {'modules': modules})

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import redirect, render
from django.forms import inlineformset_factory
from .models import Module, Question
from .forms import ModuleForm, QuestionForm

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_add_module(request):
    # Inline formset for questions (no instance yet, so use 'None')
    QuestionFormSet = inlineformset_factory(
        Module,
        Question,
        form=QuestionForm,
        extra=2,         # Show 2 blank forms for adding new questions
        can_delete=True  # Allow deleting questions (for inlines that are left blank)
    )

    if request.method == "POST":
        form = ModuleForm(request.POST)
        formset = QuestionFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            module = form.save()
            # Assign the module to each question before saving
            questions = formset.save(commit=False)
            for question in questions:
                question.module = module
                question.save()
            # Handle deletions
            for obj in formset.deleted_objects:
                obj.delete()
            return redirect("module_list")  # Change to your module list URL name
    else:
        form = ModuleForm()
        formset = QuestionFormSet()

    return render(request, "codingapp/teacher_module_form.html", {
        "form": form,
        "formset": formset,
    })

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404, redirect, render
from django.forms import inlineformset_factory
from .models import Module, Question
from .forms import ModuleForm, QuestionForm

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_edit_module(request, module_id):
    print("### USING teacher_edit_module VIEW ###")

    module = get_object_or_404(Module, id=module_id)
    
    # Inline formset for questions related to this module
    QuestionFormSet = inlineformset_factory(
        Module,
        Question,
        form=QuestionForm,
        extra=1,         # Show 1 blank form for adding a new question
        can_delete=True  # Allow deleting questions
    )
    
    if request.method == "POST":
        form = ModuleForm(request.POST, instance=module)
        formset = QuestionFormSet(request.POST, instance=module)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect("module_list")  # Change to your module list URL name
    else:
        form = ModuleForm(instance=module)
        formset = QuestionFormSet(instance=module)
    
    return render(request, "codingapp/teacher_module_form.html", {
        "form": form,
        "formset": formset,
        "module": module,
    })

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_delete_module(request, module_id):
    module = get_object_or_404(Module, id=module_id)
    if request.method == "POST":
        module.delete()
        return redirect('teacher_module_list')
    return render(request, 'codingapp/teacher_module_confirm_delete.html', {'module': module})

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_dashboard(request):
    # You can add stats, recent activity, etc.
    return render(request, 'codingapp/teacher_dashboard.html')

from .models import Question
from .forms import QuestionForm

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_question_list(request):
    questions = Question.objects.filter()
    return render(request, 'codingapp/teacher_question_list.html', {'questions': questions})


from .forms import QuestionForm, TestCaseFormSet

from .forms import QuestionForm, TestCaseFormSet
from .models import Question
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect, get_object_or_404

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_question_form(request, question_id=None):
    question = get_object_or_404(Question, id=question_id) if question_id else None

    if request.method == "POST":
        form = QuestionForm(request.POST, instance=question)
        formset = TestCaseFormSet(request.POST, prefix='testcases')
        if form.is_valid():
            q = form.save(commit=False)
            if form.cleaned_data.get('question_type') == "coding":
                if formset.is_valid():
                    tc_list = []
                    for f in formset.cleaned_data:
                        if f and not f.get('DELETE', False) and f.get('input') and f.get('expected_output'):
                            tc_list.append({
                                'input': f['input'],
                                'expected_output': [line.strip() for line in f['expected_output'].splitlines() if line.strip()],
                            })
                    q.test_cases = tc_list
                else:
                    # show errors
                    return render(request, "codingapp/teacher_question_form.html", {
                        "form": form, "formset": formset, "question": question, "action": "Edit" if question_id else "Add"
                    })
            q.save()
            form.save_m2m()
            return redirect("teacher_question_list")
    else:
        form = QuestionForm(instance=question)
        # Only for edit, prefill test cases for coding questions
        initial_test_cases = []
        if question and question.question_type == "coding":
            for tc in (question.test_cases or []):
                initial_test_cases.append({
                    'input': tc['input'],
                    'expected_output': "\n".join(tc['expected_output']),
                })
        formset = TestCaseFormSet(initial=initial_test_cases, prefix='testcases')

    return render(request, "codingapp/teacher_question_form.html", {
        "form": form,
        "formset": formset,
        "question": question,
        "action": "Edit" if question_id else "Add"
    })


@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_delete_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    if request.method == "POST":
        question.delete()
        return redirect('teacher_question_list')
    return render(request, 'codingapp/teacher_question_confirm_delete.html', {'question': question})

from .models import Assessment
from .forms import AssessmentForm  # You'll need to create this form

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_assessment_list(request):
    assessments = Assessment.objects.all()
    return render(request, 'codingapp/teacher_assessment_list.html', {'assessments': assessments})

from .forms import AssessmentForm  # Ensure AssessmentForm is imported
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test
from .models import Assessment, AssessmentQuestion # It's good practice to import models you work with


@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_add_assessment(request):
    if request.method == "POST":
        form = AssessmentForm(request.POST)
        if form.is_valid():
            # Step 1: Save the main Assessment object. This also saves simple M2M fields like 'groups'.
            assessment = form.save() 
            
            # Step 2: Manually handle the 'questions' relationship using the 'through' model.
            # This is the same logic that works correctly in your edit view.
            selected_questions = form.cleaned_data.get('questions')
            
            if selected_questions:
                for question in selected_questions:
                    AssessmentQuestion.objects.create(assessment=assessment, question=question)

            messages.success(request, 'Assessment created successfully!')
            return redirect('teacher_assessment_list')
    else:
        form = AssessmentForm()
        
    return render(request, 'codingapp/teacher_assessment_form.html', {'form': form, 'action': 'Add'})
    
@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_edit_assessment(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    if request.method == "POST":
        form = AssessmentForm(request.POST, instance=assessment)
        if form.is_valid():
            # First, save the assessment instance and its direct M2M fields (like 'groups').
            # The 'commit=False' is not strictly needed here but is good practice if you have more complex logic.
            assessment = form.save(commit=False)
            assessment.save()
            # This saves the 'groups' field correctly.
            form.save_m2m()

            # Now, manually handle the 'questions' relationship through the intermediary model.
            selected_questions = form.cleaned_data.get('questions')

            # Clear out the old question relationships for this assessment.
            assessment.assessmentquestion_set.all().delete()

            # Create the new question relationships from the selection.
            for question in selected_questions:
                AssessmentQuestion.objects.create(assessment=assessment, question=question)

            messages.success(request, "Assessment updated successfully.")
            return redirect('teacher_assessment_list')
    else:
        # For the GET request, pre-populate the form with the currently selected questions.
        initial_data = {
            'questions': assessment.assessmentquestion_set.values_list('question_id', flat=True)
        }
        form = AssessmentForm(instance=assessment, initial=initial_data)

    return render(request, 'codingapp/teacher_assessment_form.html', {
        'form': form,
        'action': 'Edit',
        'assessment': assessment
    })
    
@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_delete_assessment(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    if request.method == "POST":
        assessment.delete()
        return redirect('teacher_assessment_list')
    return render(request, 'codingapp/teacher_assessment_confirm_delete.html', {'assessment': assessment})

from .models import Group
from .forms import GroupForm

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_group_list(request):
    groups = Group.objects.all()
    return render(request, 'codingapp/teacher_group_list.html', {'groups': groups})

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_add_group(request):
    if request.method == "POST":
        form = GroupForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('teacher_group_list')
    else:
        form = GroupForm()
    return render(request, 'codingapp/teacher_group_form.html', {'form': form, 'action': 'Add'})

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_edit_group(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    if request.method == "POST":
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            return redirect('teacher_group_list')
    else:
        form = GroupForm(instance=group)
    return render(request, 'codingapp/teacher_group_form.html', {'form': form, 'action': 'Edit'})

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_delete_group(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    if request.method == "POST":
        group.delete()
        return redirect('teacher_group_list')
    return render(request, 'codingapp/teacher_group_confirm_delete.html', {'group': group})

import openpyxl
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.contrib import messages

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test
import openpyxl
from .models import UserProfile, Role, Department, Group


@user_passes_test(lambda u: hasattr(u, "userprofile") and u.userprofile.role and u.userprofile.role.name.lower() == "admin")
def bulk_user_upload(request):
    """
    Allows admin to upload users in bulk from an Excel file (.xlsx or .xls).
    Supports columns: username, full_name, email, password, role, department, group.
    Department is optional.
    """
    result = None

    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {e}")
            return redirect("admin_manage_users")

        # Expected headers (now includes optional 'department')
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        expected_headers = ["username", "full_name", "email", "password", "role", "department", "group"]

        # Validate headers (at least required ones)
        required_headers = ["username", "email", "password", "role"]
        if any(h not in headers for h in required_headers):
            messages.error(
                request,
                f"Invalid headers. Required: {', '.join(required_headers)} | "
                f"Optional: department, group"
            )
            return redirect("admin_manage_users")

        created = skipped = 0
        errors = []

        # Get index mapping for flexible column order
        col_index = {h: headers.index(h) for h in headers if h}

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                username = str(row[col_index.get("username")]).strip() if row[col_index.get("username")] else None
                email = (row[col_index.get("email")] or "").strip()
                password = (row[col_index.get("password")] or "").strip() or "password123"
                full_name = (row[col_index.get("full_name")] or "").strip()
                role_name = (row[col_index.get("role")] or "").strip()
                dept_name = (row[col_index.get("department")] or "").strip() if "department" in col_index else ""
                group_name = (row[col_index.get("group")] or "").strip() if "group" in col_index else ""

                if not username or not email or not password:
                    errors.append(f"Row {i}: Missing required fields.")
                    continue

                if User.objects.filter(username=username).exists():
                    skipped += 1
                    continue

                user = User.objects.create_user(username=username, email=email, password=password)
                profile, _ = UserProfile.objects.get_or_create(user=user)
                profile.full_name = full_name

                # Assign role
                if role_name:
                    role = Role.objects.filter(name__iexact=role_name).first()
                    if role:
                        profile.role = role

                # Assign department (optional)
                if dept_name:
                    dept = Department.objects.filter(name__iexact=dept_name).first()
                    if dept:
                        profile.department = dept

                profile.save()

                # Group assignment (optional)
                if group_name:
                    group, _ = Group.objects.get_or_create(name=group_name)
                    group.students.add(user)

                created += 1

            except Exception as e:
                errors.append(f"Row {i}: {e}")

        result = {"created": created, "skipped": skipped, "errors": errors}

        if created > 0:
            messages.success(request, f"✅ {created} users added successfully. {skipped} skipped.")
        if errors:
            messages.warning(request, f"⚠️ Some rows failed: {len(errors)} issues.")
            for err in errors[:5]:  # show only top 5 errors
                messages.info(request, err)

        return redirect("admin_manage_users")

    return render(request, "codingapp/bulk_user_upload.html", {"result": result})



from django.contrib.auth.decorators import user_passes_test, login_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Question, Submission

def is_admin_or_teacher(user):
    return user.is_superuser or user.groups.filter(name='Teachers').exists()

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Assessment, AssessmentSubmission, AssessmentSession

def is_admin_or_teacher(user):
    return user.is_superuser or user.groups.filter(name='Teachers').exists()

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Assessment, AssessmentSubmission, AssessmentSession

def is_admin_or_teacher(user):
    return user.is_superuser or user.groups.filter(name='Teachers').exists()

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Assessment, AssessmentSubmission, AssessmentSession

def is_admin_or_teacher(user):
    return user.is_superuser or user.groups.filter(name='Teachers').exists()

@login_required
@user_passes_test(is_admin_or_teacher)
def reset_submissions_admin(request):
    users = User.objects.all().order_by('username')
    assessments = Assessment.objects.all().order_by('title')
    selected_user = None
    selected_assessment = None

    if request.method == "POST":
        user_id = request.POST.get("user_id")
        assessment_id = request.POST.get("assessment_id")
        selected_user = get_object_or_404(User, pk=user_id)
        selected_assessment = get_object_or_404(Assessment, pk=assessment_id)

        # Delete submissions & session
        AssessmentSubmission.objects.filter(user=selected_user, assessment=selected_assessment).delete()
        AssessmentSession.objects.filter(user=selected_user, assessment=selected_assessment).delete()
        messages.success(request, f"Reset all submissions and session for {selected_user.username} in '{selected_assessment.title}'.")
        return redirect('reset_submissions_admin')

    return render(request, 'codingapp/reset_submissions_admin.html', {
        'users': users,
        'assessments': assessments,
        'selected_user': selected_user,
        'selected_assessment': selected_assessment,
    })

from django.http import JsonResponse

def clear_splash_flag(request):
    request.session.pop('force_splash', None)
    return JsonResponse({'cleared': True})

import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from .models import Question


@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_bulk_upload_mcq(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file)
            created = 0
            for _, row in df.iterrows():
                options = [row[f'Option{i}'] for i in range(1, 5) if pd.notnull(row.get(f'Option{i}'))]
                Question.objects.create(
                    title=row['Question Text'][:200],
                    description=row.get('Description', ''),
                    question_type='mcq',
                    options=options,
                    correct_answer=row['Correct Answer'],
                )
                created += 1
            messages.success(request, f"{created} MCQ questions imported successfully!")
        except Exception as e:
            messages.error(request, f"Error processing file: {e}")
    return render(request, 'codingapp/teacher_bulk_upload_mcq.html')


from .models import Quiz, QuizSubmission, QuizAnswer
from django.utils import timezone
import random
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from .models import Quiz, QuizSubmission, QuizAnswer

@login_required
def take_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    session_key = f'quiz_{quiz_id}_question_order'

    if request.method == 'POST':
        # Retrieve shuffled order from session
        question_ids = request.session.get(session_key)
        if not question_ids:
            messages.error(request, "Session expired. Please retake the quiz.")
            return redirect('take_quiz', quiz_id=quiz.id)
        questions = list(quiz.questions.filter(id__in=question_ids))
        # Preserve order
        questions.sort(key=lambda q: question_ids.index(q.id))
        # Process answers
        submission = QuizSubmission.objects.create(user=request.user, quiz=quiz, submitted_at=timezone.now())
        score = 0
        for q in questions:
            selected = request.POST.get(f'question_{q.id}')
            if selected:
                QuizAnswer.objects.create(submission=submission, question=q, selected_option=selected)
                if selected == q.correct_answer:
                    score += 1
        submission.score = score
        submission.save()
        # Remove order from session
        del request.session[session_key]
        messages.success(request, f"You scored {score} out of {len(questions)}")
        return redirect('quiz_result', submission_id=submission.id)
    else:
        # First load: shuffle and store order in session
        questions = list(quiz.questions.all())
        random.shuffle(questions)
        request.session[session_key] = [q.id for q in questions]
    return render(request, 'codingapp/take_quiz.html', {'quiz': quiz, 'questions': questions})

from django.contrib.auth.decorators import user_passes_test, login_required
from django.shortcuts import render, redirect, get_object_or_404
from .models import Quiz
from .forms import QuizForm

from .models import Quiz
from .forms import QuizForm
from django.contrib.auth.decorators import user_passes_test

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_quiz_list(request):
    quizzes = Quiz.objects.filter(created_by=request.user)
    return render(request, 'codingapp/teacher_quiz_list.html', {'quizzes': quizzes})

# codingapp/views.py (add at top if not already present)
import csv
import io
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.core.files.uploadedfile import UploadedFile

# openpyxl will be used for .xlsx parsing
try:
    import openpyxl
except Exception:
    openpyxl = None

from .models import Quiz, Question, Module
from .forms import QuizForm
from .utils import is_teacher  # adjust to your is_teacher import/location


@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_quiz_create(request, quiz_id=None):
    """
    Handles creating/editing a Quiz and bulk-uploading MCQs.
    If 'upload_questions' is present in POST, we process the uploaded file and return a result dict to the template.
    If not, we treat it as a normal Save Quiz action and redirect to quiz list.
    """
    # If editing existing quiz (optional), load it
    quiz = None
    if quiz_id:
        quiz = get_object_or_404(Quiz, id=quiz_id, created_by=request.user)

    if request.method == "POST":
        form = QuizForm(request.POST, instance=quiz)
        upload_flag = 'upload_questions' in request.POST
        uploaded_file = request.FILES.get('bulk_file')  # name from template

        if not form.is_valid():
            # show form errors and do not proceed
            return render(request, 'codingapp/teacher_quiz_form.html', {
                'form': form,
                'action': 'Edit' if quiz else 'Create',
            })

        # Save (create or update) the quiz first so we can attach Qs to it
        quiz = form.save(commit=False)
        quiz.created_by = request.user
        quiz.save()
        form.save_m2m()

        # If user clicked Upload Questions, parse file and attach questions, but stay on same page and show summary
        if upload_flag:
            result = {'created': 0, 'skipped': 0, 'errors': []}
            if not uploaded_file:
                result['errors'].append("No file uploaded. Please choose a .xlsx or .csv file.")
                return render(request, 'codingapp/teacher_quiz_form.html', {
                    'form': form,
                    'action': 'Edit' if quiz else 'Create',
                    'result': result
                })

            # Basic safety check: file size limit (example 5 MB)
            max_size = 5 * 1024 * 1024
            if isinstance(uploaded_file, UploadedFile) and uploaded_file.size > max_size:
                result['errors'].append(f"File too large. Max allowed is {max_size // (1024*1024)} MB.")
                return render(request, 'codingapp/teacher_quiz_form.html', {
                    'form': form,
                    'action': 'Edit' if quiz else 'Create',
                    'result': result
                })

            # Read and parse file
            try:
                # xlsx
                if uploaded_file.name.lower().endswith(('.xlsx', '.xls')) and openpyxl:
                    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        result['errors'].append("Spreadsheet is empty.")
                        return render(request, 'codingapp/teacher_quiz_form.html', {'form': form, 'action': 'Create', 'result': result})

                    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
                    data_rows = rows[1:]
                    reader_rows = []
                    for r in data_rows:
                        # map header -> cell text
                        rowdict = {headers[i]: (str(r[i]).strip() if r[i] is not None else "") for i in range(len(headers))}
                        reader_rows.append(rowdict)

                else:
                    # assume CSV (read as text)
                    raw = uploaded_file.read()
                    # If it's bytes, decode; some UploadedFile may already give bytes
                    if isinstance(raw, bytes):
                        try:
                            text = raw.decode('utf-8')
                        except UnicodeDecodeError:
                            try:
                                text = raw.decode('latin-1')
                            except Exception:
                                result['errors'].append("Unable to decode CSV file. Use UTF-8 or Latin-1 encoding.")
                                return render(request, 'codingapp/teacher_quiz_form.html', {'form': form, 'action': 'Create', 'result': result})
                    else:
                        text = str(raw)

                    f = io.StringIO(text)
                    reader = csv.DictReader(f)
                    reader_rows = []
                    for row in reader:
                        # normalize keys to lowercase
                        row = {str(k).strip().lower(): (v.strip() if v is not None else "") for k, v in row.items()}
                        reader_rows.append(row)
            except Exception as e:
                result['errors'].append(f"Failed to parse file: {e}")
                return render(request, 'codingapp/teacher_quiz_form.html', {'form': form, 'action': 'Create', 'result': result})

            # Now iterate rows and create Question objects
            with transaction.atomic():
                for idx, row in enumerate(reader_rows, start=1):
                    try:
                        # support multiple possible header names
                        title = (row.get('question text') or row.get('title') or row.get('question') or "").strip()
                        if not title:
                            result['errors'].append(f"Row {idx}: missing question text/title.")
                            continue

                        # Collect options from Option1..Option10 or opt1..opt4
                        options = []
                        for i in range(1, 11):
                            k = f"option{i}"
                            if k in row and row[k]:
                                options.append(row[k])
                        if not options:
                            # try "option 1" with space or Option1 capitalized handled by lowercase keys earlier
                            for k in ('option1','option2','option3','option4','a','b','c','d'):
                                if row.get(k):
                                    options.append(row.get(k))

                        # remove empty options and ensure at least two
                        options = [o for o in options if o and o.strip()]
                        if len(options) < 2:
                            result['errors'].append(f"Row {idx}: need at least two non-empty options.")
                            continue

                        # correct answer
                        corr = (row.get('correct answer') or row.get('correct_answer') or
                                row.get('correct') or row.get('answer') or "").strip()
                        if not corr:
                            result['errors'].append(f"Row {idx}: missing Correct Answer.")
                            continue

                        # Determine correct_answer text
                        correct_text = None
                        if corr.isdigit():
                            ix = int(corr) - 1
                            if 0 <= ix < len(options):
                                correct_text = options[ix]
                            else:
                                result['errors'].append(f"Row {idx}: correct answer index {corr} out of range.")
                                continue
                        else:
                            # match ignoring case
                            matched = None
                            for opt in options:
                                if opt.strip().lower() == corr.strip().lower():
                                    matched = opt
                                    break
                            if matched:
                                correct_text = matched
                            else:
                                # not found — still accept but warn
                                correct_text = corr
                                result['errors'].append(f"Row {idx}: correct answer '{corr}' not found among options; saved as-is.")

                        description = row.get('description') or row.get('explanation') or ""

                        # Optional: module
                        module_obj = None
                        module_key = row.get('module') or row.get('module name') or row.get('module_slug')
                        if module_key:
                            module_obj = Module.objects.filter(name__iexact=module_key).first() \
                                         or Module.objects.filter(slug__iexact=module_key).first()
                            if not module_obj:
                                # do not fail; just warn
                                result['errors'].append(f"Row {idx}: module '{module_key}' not found (question will be created without module).")

                        # Check duplicates: same title + same module => skip
                        duplicate_qs = Question.objects.filter(title__iexact=title)
                        if module_obj:
                            duplicate_qs = duplicate_qs.filter(module=module_obj)
                        if duplicate_qs.exists():
                            result['skipped'] += 1
                            continue

                        # Create question
                        q = Question.objects.create(
                            question_type='mcq',
                            title=title,
                            description=description,
                            options=options,
                            correct_answer=correct_text,
                            module=module_obj
                        )
                        quiz.questions.add(q)
                        result['created'] += 1

                    except Exception as e:
                        result['errors'].append(f"Row {idx}: Unexpected error: {e}")
                        # continue with other rows

            # render same page with result summary
            return render(request, 'codingapp/teacher_quiz_form.html', {
                'form': form,
                'action': 'Edit' if quiz else 'Create',
                'result': result
            })

        # else: not upload_flag -> regular Save Quiz flow (redirect to list)
        messages.success(request, "Quiz saved.")
        return redirect('teacher_quiz_list')

    else:
        form = QuizForm(instance=quiz)

    return render(request, 'codingapp/teacher_quiz_form.html', {
        'form': form,
        'action': 'Edit' if quiz else 'Create'
    })


@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_quiz_edit(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, created_by=request.user)
    if request.method == "POST":
        form = QuizForm(request.POST, instance=quiz)
        if form.is_valid():
            form.save()
            return redirect('teacher_quiz_list')
    else:
        form = QuizForm(instance=quiz)
    return render(request, 'codingapp/teacher_quiz_form.html', {'form': form, 'action': 'Edit'})

@user_passes_test(is_teacher, login_url='/dashboard/')
def teacher_quiz_delete(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id, created_by=request.user)
    if request.method == "POST":
        quiz.delete()
        return redirect('teacher_quiz_list')
    return render(request, 'codingapp/teacher_quiz_confirm_delete.html', {'quiz': quiz})

@login_required
def quiz_result(request, submission_id):
    submission = get_object_or_404(QuizSubmission, id=submission_id, user=request.user)
    answers = submission.answers.select_related('question')
    return render(request, 'codingapp/quiz_result.html', {'submission': submission, 'answers': answers})

from django.contrib.auth.decorators import user_passes_test, login_required
from django.shortcuts import render, get_object_or_404
from .models import Assessment, AssessmentSession, Quiz, QuizSubmission

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from .models import Assessment, QuizSubmission

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from .models import Assessment, QuizSubmission

@login_required
def quiz_leaderboard(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    quiz = assessment.quiz  # assumes Assessment has a ForeignKey to Quiz
    submissions = QuizSubmission.objects.filter(quiz=quiz).select_related('user').order_by('-score', 'submitted_at')
    print("Quiz ID for leaderboard:", quiz.id)
    print("QuizSubmission count:", QuizSubmission.objects.filter(quiz=quiz).count())

    return render(request, 'codingapp/quiz_leaderboard.html', {
        'assessment': assessment,
        'quiz': quiz,
        'submissions': submissions,
    })

from .models import Note
from .forms import NoteForm
from django.contrib.auth.decorators import login_required


from django.db.models import Q
from .models import Note

from codingapp.utils import is_teacher, is_hod, is_admin  # make sure this import exists


@login_required
def notes_list(request):
    """
    Students see only notes from their groups.
    Teachers/HOD/Admins see all notes from their department.
    """
    accessible_groups = get_user_accessible_groups(request.user)

    # Teachers, HODs, Admins can see all notes for groups they manage
    notes = Note.objects.filter(
        Q(group__in=accessible_groups) | Q(group__isnull=True)
    ).distinct()

    can_manage_notes = is_teacher(request.user) or is_hod(request.user) or is_admin(request.user)

    return render(request, "codingapp/notes_list.html", {
        "notes": notes,
        "can_manage_notes": can_manage_notes,
    })


@login_required
def add_note(request):
    if not (is_teacher(request.user) or is_hod(request.user) or is_admin(request.user)):
        messages.error(request, "You do not have permission to upload study materials.")
        return redirect('notes_list')

    form = NoteForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        note = form.save(commit=False)
        note.uploaded_by = request.user
        note.save()
        messages.success(request, "Note uploaded successfully.")
        return redirect('notes_list')

    return render(request, "codingapp/add_note.html", {'form': form})


@login_required
def edit_note(request, note_id):
    note = get_object_or_404(Note, id=note_id)

    if not (is_teacher(request.user) or is_hod(request.user) or is_admin(request.user)) or note.uploaded_by != request.user:
        messages.error(request, "You cannot edit this note.")
        return redirect('notes_list')

    form = NoteForm(request.POST or None, request.FILES or None, instance=note)
    if form.is_valid():
        form.save()
        messages.success(request, "Note updated successfully.")
        return redirect('notes_list')

    return render(request, "codingapp/add_note.html", {'form': form, 'edit_mode': True})


@login_required
def delete_note(request, note_id):
    note = get_object_or_404(Note, id=note_id)

    if not (is_teacher(request.user) or is_hod(request.user) or is_admin(request.user)) or note.uploaded_by != request.user:
        messages.error(request, "You cannot delete this note.")
        return redirect('notes_list')

    if request.method == 'POST':
        note.delete()
        messages.success(request, "Note deleted successfully.")
        return redirect('notes_list')

    return render(request, "codingapp/confirm_delete_note.html", {'note': note})


# List all notices for current user (group/for_everyone)
@login_required
def notice_list(request):
    user = request.user
    notices = Notice.objects.filter(
        Q(for_everyone=True) | Q(group__in=user.custom_groups.all())
    ).order_by('-created_at').distinct()
    # Unread set
    unread_ids = set(
        NoticeReadStatus.objects.filter(
            user=user, is_read=False, notice__in=notices
        ).values_list('notice_id', flat=True)
    )
    return render(request, "codingapp/notice_list.html", {"notices": notices, "unread_ids": unread_ids})


from django.db.models import Q

# Detail view and mark as read
@login_required
def notice_detail(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    # Only show if for group/everyone
    if not notice.for_everyone and notice.group not in request.user.custom_groups.all():
        return render(request, "codingapp/permission_denied.html", status=403)
    # Mark as read
    NoticeReadStatus.objects.get_or_create(
        user=request.user, notice=notice,
        defaults={"is_read": True, "read_at": timezone.now()}
    )
    return render(request, "codingapp/notice_detail.html", {"notice": notice})

from django.db.models import Q

# Staff: Add notice
from codingapp.utils import is_teacher, is_hod, is_admin

@login_required
def add_notice(request):
    if request.method == "POST":
        form = NoticeForm(request.POST, request.FILES)
        if form.is_valid():
            notice = form.save(commit=False)
            notice.created_by = request.user
            notice.save()
            form.save_m2m()
            # Pre-create unread status for all targeted users
            users = User.objects.filter(is_active=True)
            if not notice.for_everyone:
                users = users.filter(custom_groups=notice.group)
            for user in users:
                NoticeReadStatus.objects.get_or_create(user=user, notice=notice)
            return redirect("notice_list")
    else:
        form = NoticeForm()
    return render(request, "codingapp/add_notice.html", {"form": form})

# Staff: Edit and Delete Notice (similar pattern, optional for now)
from django.contrib import messages
from codingapp.utils import is_teacher, is_hod, is_admin

@login_required
def edit_notice(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    if request.method == "POST":
        form = NoticeForm(request.POST, request.FILES, instance=notice)
        if form.is_valid():
            form.save()
            messages.success(request, "Notice updated successfully.")
            return redirect("notice_detail", pk=notice.pk)
    else:
        form = NoticeForm(instance=notice)
    return render(request, "codingapp/edit_notice.html", {"form": form, "notice": notice})
from codingapp.utils import is_teacher, is_hod, is_admin
from codingapp.utils import is_teacher, is_hod, is_admin

@login_required
def delete_notice(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    if request.method == "POST":
        notice.delete()
        messages.success(request, "Notice deleted.")
        return redirect("notice_list")
    return render(request, "codingapp/delete_notice.html", {"notice": notice})

from .forms import UserEditForm, UserProfileEditForm
from django.contrib.auth.decorators import login_required

@login_required
def edit_profile(request):
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    if request.method == "POST":
        uform = UserEditForm(request.POST, instance=user)
        pform = UserProfileEditForm(request.POST, request.FILES, instance=profile)
        if uform.is_valid() and pform.is_valid():
            uform.save()
            pform.save()
            return redirect("dashboard")
    else:
        uform = UserEditForm(instance=user)
        pform = UserProfileEditForm(instance=profile)
    return render(request, "codingapp/edit_profile.html", {
        "uform": uform,
        "pform": pform,
    })

import openpyxl
from django.contrib.auth.decorators import user_passes_test
from .forms import BulkMCQUploadForm
from .models import Question
from codingapp.utils import is_teacher, is_hod, is_admin

@login_required
def bulk_mcq_upload(request):
    result = None
    if request.method == "POST":
        form = BulkMCQUploadForm(request.POST, request.FILES)
        if form.is_valid():
            module = form.cleaned_data['module']
            excel_file = form.cleaned_data['file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            expected = ["Question Text", "Option1", "Option2", "Option3", "Option4", "Correct Answer", "Description"]
            # Header validation
            if any(h not in headers for h in expected):
                result = {"created": 0, "skipped": 0, "errors": [f"Invalid headers. Expected: {', '.join(expected)}"]}
            else:
                created = skipped = 0
                errors = []
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    row_data = dict(zip(headers, row))
                    title = str(row_data.get("Question Text") or "").strip()
                    options = [str(row_data.get(f"Option{n}") or "").strip() for n in range(1, 5) if row_data.get(f"Option{n}") is not None]
                    correct_answer = str(row_data.get("Correct Answer") or "").strip()
                    description = str(row_data.get("Description") or "").strip()
                    # Validation
                    if not title or len(options) != 4 or not correct_answer:
                        errors.append(f"Row {i}: Missing question/options/correct answer.")
                        continue
                    # Answer match validation (case-insensitive, strip spaces)
                    if not any(correct_answer.strip().lower() == opt.strip().lower() for opt in options):
                        errors.append(f"Row {i}: Correct answer '{correct_answer}' not among options.")
                        continue
                    # Skip duplicate
                    if Question.objects.filter(module=module, title__iexact=title).exists():
                        skipped += 1
                        continue
                    try:
                        # Store options as JSON list
                        question = Question.objects.create(
                            module=module,
                            title=title,
                            description=description,
                            question_type="mcq",
                            options=options,
                            correct_answer=correct_answer,
                        )
                        created += 1
                    except Exception as e:
                        errors.append(f"Row {i}: {e}")
                result = {"created": created, "skipped": skipped, "errors": errors}
    else:
        form = BulkMCQUploadForm()
    return render(request, "codingapp/bulk_mcq_upload.html", {"form": form, "result": result})


from django.shortcuts import render, redirect, get_object_or_404
from .models import Course
from .forms import CourseForm, CourseContentFormSet
from django.contrib.auth.decorators import login_required, user_passes_test

from django.db.models import Q
from codingapp.models import Course, Group  # ✅ Make sure this is your model
from django.contrib.auth.decorators import login_required

@login_required
def course_list(request):
    role = request.user.userprofile.role.name.lower()

    if role == "admin":
        courses = Course.objects.all()
    else:
        user_groups = Group.objects.filter(students=request.user)
        courses = Course.objects.filter(
            Q(is_public=True) |
            Q(groups__in=user_groups) |
            Q(created_by=request.user)
        ).distinct()

    return render(
        request,
        'codingapp/courses/list.html',
        {'courses': courses}
    )


@login_required
def course_detail(request, pk):
    from django.shortcuts import get_object_or_404, render
    from codingapp.models import Course

    course = get_object_or_404(Course, pk=pk)

    denied = deny_access_if_not_allowed(request, course)
    if denied:
        return denied

    user_profile = request.user.userprofile
    role = user_profile.role.name.lower()

    # -----------------------------
    # ADMIN → always allowed
    # -----------------------------
    if role == "admin":
        pass

    # -----------------------------
    # NON-ADMIN ACCESS CHECK
    # -----------------------------
    else:
        if (
            not course.is_public and
            not course.groups.filter(
                id__in=request.user.custom_groups.values_list('id', flat=True)
            ).exists()
        ):
            return render(request, 'codingapp/courses/denied.html')

    lang_list = ['python', 'c', 'cpp', 'java', 'javascript']

    return render(
        request,
        'codingapp/courses/detail.html',
        {
            'course': course,
            'lang_list': lang_list
        }
    )




@login_required
@user_passes_test(is_teacher_or_admin, login_url='/dashboard/')
def create_course(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        formset = CourseContentFormSet(request.POST, prefix='contents')

        if form.is_valid() and formset.is_valid():
            course = form.save(commit=False)
            course.created_by = request.user
            course.save()
            form.save_m2m()

            formset.instance = course
            formset.save()

            messages.success(request, "Course created successfully.")
            return redirect('manage_courses')
        else:
            messages.error(request, "Please correct the errors.")
    else:
        form = CourseForm()
        formset = CourseContentFormSet(prefix='contents')

    return render(
        request,
        'codingapp/courses/create.html',
        {
            'form': form,
            'formset': formset,
        }
    )


from django.contrib import messages

@login_required
@user_passes_test(is_teacher_or_admin, login_url='/dashboard/')
def manage_courses(request):
    role = request.user.userprofile.role.name.lower()

    if role == "admin":
        courses = Course.objects.all()
    else:
        courses = Course.objects.filter(created_by=request.user)

    return render(
        request,
        'codingapp/courses/manage.html',
        {'courses': courses}
    )


@login_required
@user_passes_test(is_teacher_or_admin, login_url='/dashboard/')
def edit_course(request, pk):
    role = request.user.userprofile.role.name.lower()

    if role == "admin":
        course = get_object_or_404(Course, pk=pk)
    else:
        course = get_object_or_404(Course, pk=pk, created_by=request.user)

    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        formset = CourseContentFormSet(request.POST, instance=course)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "Course updated successfully.")
            return redirect('manage_courses')
        else:
            messages.error(request, "Please correct the errors.")
    else:
        form = CourseForm(instance=course)
        formset = CourseContentFormSet(instance=course)

    return render(
        request,
        'codingapp/courses/create.html',
        {
            'form': form,
            'formset': formset,
        }
    )


@login_required
@user_passes_test(is_teacher_or_admin, login_url='/dashboard/')
def delete_course(request, pk):
    role = request.user.userprofile.role.name.lower()

    if role == "admin":
        course = get_object_or_404(Course, pk=pk)
    else:
        course = get_object_or_404(Course, pk=pk, created_by=request.user)

    if request.method == 'POST':
        course.delete()
        messages.success(request, "Course deleted.")
        return redirect('manage_courses')

    return render(
        request,
        'codingapp/courses/delete_confirm.html',
        {'course': course}
    )

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.core.files.storage import default_storage

@csrf_exempt
def ckeditor_upload(request):
    if request.method == "POST" and request.FILES.get("upload"):
        file = request.FILES["upload"]
        file_path = default_storage.save(f"ckeditor_uploads/{file.name}", file)
        file_url = default_storage.url(file_path)
        return JsonResponse({
            "url": file_url
        })
    return JsonResponse({"error": "Invalid upload"}, status=400)


from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.http import JsonResponse
import json, requests

@csrf_exempt
@login_required
def run_code_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            code = data.get("code", "")
            language = data.get("language", "python")
            stdin = data.get("stdin", "")  # ✅ Custom input from user

            if not code or not language:
                return JsonResponse({"error": "Missing code or language."}, status=400)

            payload = {
                "language": language,
                "version": "*",
                "files": [{"name": "main", "content": code}],
                "stdin": stdin  # ✅ Send custom input
            }

            response = requests.post(settings.PISTON_API_URL, json=payload, timeout=10)
            response.raise_for_status()
            result = response.json()

            return JsonResponse({
                "output": result.get("run", {}).get("stdout", ""),
                "error": result.get("run", {}).get("stderr", "")
            })

        except requests.exceptions.RequestException as req_err:
            return JsonResponse({"error": f"Piston API error: {str(req_err)}"}, status=502)
        except Exception as e:
            return JsonResponse({"error": f"Server error: {str(e)}"}, status=500)

    return JsonResponse({"error": "Only POST method allowed."}, status=405)




from django.db.models import Avg, Q
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from django.contrib.auth.models import User
from codingapp.models import Submission, QuizSubmission, Module, ModuleCompletion, Group, UserProfile
from codingapp.utils import get_user_accessible_groups
from codingapp.utils import is_teacher, is_hod, is_admin
from codingapp.utils import get_user_accessible_groups
from codingapp.utils import (
    get_visible_students,
    get_student_performance
)


@login_required
def student_performance_list(request):
    from codingapp.utils import get_student_performance
    from codingapp.models import Group, Department

    profile = request.user.userprofile
    role = profile.role.name.lower()

    # -------------------------------
    # TEACHER → Groups assigned
    # -------------------------------
    if role == "teacher":
        groups = (
            Group.objects
            .filter(teachers=request.user)
            .prefetch_related("students")
        )

        data = []
        for group in groups:
            data.append({
                "group": group,
                "students": [
                    get_student_performance(s)
                    for s in group.students.all()
                ]
            })

        return render(
            request,
            "codingapp/performance_teacher.html",
            {"groups": data}
        )

    # -------------------------------
    # HOD → All groups of department
    # -------------------------------
    if role == "hod":
        groups = (
            Group.objects
            .filter(department=profile.department)
            .prefetch_related("students")
        )

        data = []
        for group in groups:
            data.append({
                "group": group,
                "students": [
                    get_student_performance(s)
                    for s in group.students.all()
                ]
            })

        return render(
            request,
            "codingapp/performance_hod.html",
            {"groups": data}
        )

    # -------------------------------
    # ADMIN → Department → Group → Students
    # -------------------------------
    if role == "admin":
        from codingapp.models import Department

        departments = Department.objects.prefetch_related(
            "groups__students"
        )

        dept_data = []

        for dept in departments:
            group_blocks = []

            for group in dept.groups.all():  # 👈 FIXED HERE
                group_blocks.append({
                    "group": group,
                    "students": [
                        get_student_performance(student)
                        for student in group.students.all()
                    ]
                })

            dept_data.append({
                "department": dept,
                "groups": group_blocks
            })

        return render(
            request,
            "codingapp/performance_admin.html",
            {"departments": dept_data}
        )


    return render(request, "codingapp/permission_denied.html", status=403)


from django.http import HttpResponse
from django.contrib.auth.models import User
from django.db.models import Count
from .models import Submission, QuizSubmission, Module
from django.contrib.auth.decorators import login_required, user_passes_test


from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.db.models import Q
from codingapp.models import (
    Submission, QuizSubmission, Module, ModuleCompletion,
    Group, UserProfile, Assessment
)
from codingapp.utils import get_user_accessible_groups

from codingapp.utils import is_teacher, is_hod, is_admin
from codingapp.utils import get_user_accessible_groups

@login_required
def student_performance_detail(request, student_id):
    from django.shortcuts import get_object_or_404, render
    from django.contrib.auth import get_user_model
    from codingapp.utils import get_student_performance

    User = get_user_model()
    viewer_profile = request.user.userprofile
    role = viewer_profile.role.name.lower()

    # -------------------------------
    # ADMIN → any student
    # -------------------------------
    if role == "admin":
        student = get_object_or_404(User, id=student_id)

    # -------------------------------
    # HOD → students of department
    # -------------------------------
    elif role == "hod":
        student = get_object_or_404(
            User,
            id=student_id,
            custom_groups__department=viewer_profile.department
        )

    # -------------------------------
    # TEACHER → students of own groups
    # -------------------------------
    elif role == "teacher":
        student = get_object_or_404(
            User,
            id=student_id,
            custom_groups__teachers=request.user
        )

    # -------------------------------
    # STUDENT → only self
    # -------------------------------
    else:
        if request.user.id != student_id:
            return render(
                request,
                "codingapp/permission_denied.html",
                status=403
            )
        student = request.user

    # ✅ SINGLE SOURCE OF TRUTH
    performance = get_student_performance(student)

    # 🔍 Debug (keep temporarily)
    import pprint
    print("PERFORMANCE DEBUG FOR:", student.username)
    pprint.pprint(performance)

    return render(
        request,
        "codingapp/student_performance_detail.html",
        {
            "student": student,
            "performance": performance
        }
    )


@login_required
def sync_student_external_profiles(request, student_id):
    from django.shortcuts import redirect, get_object_or_404
    from django.contrib import messages
    from django.contrib.auth import get_user_model
    from codingapp.models import ExternalProfile
    from codingapp.external_services.codeforces import fetch_codeforces_stats
    from codingapp.external_services.leetcode import fetch_leetcode_stats
    from codingapp.external_services.codechef import fetch_codechef_stats
    from django.utils import timezone

    User = get_user_model()

    # Only teachers / HOD / admin
    profile = request.user.userprofile
    role = profile.role.name.lower()

    if role not in ("teacher", "hod", "admin"):
        messages.error(request, "Permission denied")
        return redirect("student_performance_detail", student_id=student_id)

    student = get_object_or_404(User, id=student_id)

    ext_profile, _ = ExternalProfile.objects.get_or_create(user=student)

    # ---------------- CODEFORCES ----------------
    if ext_profile.codeforces_username:
        cf = fetch_codeforces_stats(ext_profile.codeforces_username)
        if cf:
            ext_profile.codeforces_stats = cf

    # ---------------- LEETCODE ------------------
    if ext_profile.leetcode_username:
        lc = fetch_leetcode_stats(ext_profile.leetcode_username)
        if lc:
            ext_profile.leetcode_stats = lc

    # ---------------- CODECHEF ------------------
    if ext_profile.codechef_username:
        cc = fetch_codechef_stats(ext_profile.codechef_username)
        if cc:
            ext_profile.codechef_stats = cc

    # ---------------- HACKERRANK (MISSING PART) ----------------
    if ext_profile.hackerrank_username:
        external["hackerrank"] = {
            "username": ext_profile.hackerrank_username,
            "profile_url": f"https://www.hackerrank.com/{ext_profile.hackerrank_username}"
        }

    ext_profile.last_synced = timezone.now()
    ext_profile.save()

    messages.success(
        request,
        "External profiles synced successfully."
    )

    return redirect(
        "student_performance_detail",
        student_id=student_id
    )



import csv
from django.http import HttpResponse
from django.db.models import Avg

@login_required
@user_passes_test(is_teacher, login_url='/dashboard/')
def export_student_performance(request):
    selected_group_id = request.GET.get('group')
    search_query = request.GET.get('q', '')

    students = User.objects.filter(is_staff=False)
    if selected_group_id:
        students = students.filter(groups__id=selected_group_id)
    if search_query:
        students = students.filter(username__icontains=search_query)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="student_performance.csv"'

    writer = csv.writer(response)
    writer.writerow(['Username', 'Total Coding Submissions', 'Accepted Submissions', 'Total Quizzes', 'Average Quiz Score'])

    for student in students:
        coding_submissions = Submission.objects.filter(user=student)
        quiz_submissions = QuizSubmission.objects.filter(user=student)

        total_coding = coding_submissions.count()
        accepted_coding = coding_submissions.filter(status='Accepted').count()
        total_quizzes = quiz_submissions.count()
        avg_quiz_score = quiz_submissions.aggregate(avg=Avg('score'))['avg'] or 0

        writer.writerow([
            student.username,
            total_coding,
            accepted_coding,
            total_quizzes,
            round(avg_quiz_score, 2)
        ])

    return response

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from difflib import SequenceMatcher
from django.db.models import Max

from .models import Assessment, AssessmentSession, AssessmentSubmission, QuizSubmission
from .utils import normalize_code, penalty_factor_from_plagiarism  # ensure these exist


@login_required
def assessment_result(request, assessment_id):
    """
    Render final result for the current user's assessment session.
    - Records session end_time if not already set.
    - Computes per-question and overall plagiarism (using normalized code).
    - Computes raw totals (quiz + coding raw marks) and applies assessment-level penalty.
    - Persists penalty info to AssessmentSession if fields exist.
    """
    assessment = get_object_or_404(Assessment, id=assessment_id)
    session = get_object_or_404(AssessmentSession, user=request.user, assessment=assessment)

    # --- Ensure session end_time is recorded once (finalization) ---
    if not session.end_time:
        session.end_time = timezone.now()
        session.save(update_fields=["end_time"] if hasattr(session, "end_time") else None)

    # --- Load user's submissions for this assessment ---
    submissions = AssessmentSubmission.objects.filter(user=request.user, assessment=assessment).select_related("question")

    # Compute coding totals: raw (before per-question penalty) and penalized per-question scores
    coding_raw_total = 0.0   # sum of raw_score (or score if raw_score missing)
    coding_penalized_total = 0.0  # sum of stored penalized per-question score (sub.score)

    # ---------- PLAGIARISM: per-question and overall ----------
    per_question_plag = {}  # question_id -> max similarity %
    overall_max = 0.0

    # Preload other submissions grouped by question to avoid many DB hits
    # Build a dict: question_id -> list of other codes (strings)
    question_ids = {sub.question_id for sub in submissions}
    others_by_q = {}
    for qid in question_ids:
        others = AssessmentSubmission.objects.filter(
            assessment=assessment, question_id=qid
        ).exclude(user=request.user).values_list("code", flat=True)
        # store as list of normalized strings (skip empty)
        others_by_q[qid] = [normalize_code((c or ""), "") for c in others if (c and c.strip())]

    for sub in submissions:
        # get raw_score if available, else fallback to score
        raw = getattr(sub, "raw_score", None)
        if raw is None:
            # if raw_score not available, best-effort fallback: if sub.score == 5 assume raw 5 else 0
            raw = float(sub.score) if (hasattr(sub, "score") and sub.score is not None) else 0.0
        coding_raw_total += float(raw or 0.0)

        # penalized per-question (score) stored in sub.score
        coding_penalized_total += float(getattr(sub, "score", 0.0) or 0.0)

        my_code = (sub.code or "").strip()
        if not my_code:
            per_question_plag[sub.question_id] = 0.0
            continue

        norm_my = normalize_code(my_code, "")
        max_sim = 0.0
        for o_norm in others_by_q.get(sub.question_id, []):
            if not o_norm:
                continue
            try:
                sim = SequenceMatcher(None, norm_my, o_norm).ratio()
                if sim > max_sim:
                    max_sim = sim
            except Exception:
                continue

        pct = round(max_sim * 100, 2)
        per_question_plag[sub.question_id] = pct
        if pct > overall_max:
            overall_max = pct

    # overall average across coding questions (if any)
    overall_avg = 0.0
    if per_question_plag:
        overall_avg = round(sum(per_question_plag.values()) / len(per_question_plag), 2)

    # --- Quiz score (best attempt if multiple) ---
    quiz_score = 0.0
    if getattr(assessment, "quiz", None):
        best_quiz = QuizSubmission.objects.filter(quiz=assessment.quiz, user=request.user).order_by('-score').first()
        if best_quiz:
            quiz_score = float(best_quiz.score or 0.0)

    # --- Raw total (quiz + coding raw) and penalized total (after assessment-level penalty) ---
    raw_total = round(quiz_score + coding_raw_total, 2)
    # compute assessment-level penalty factor from overall_max (max plagiarism across their coding answers)
    factor = penalty_factor_from_plagiarism(overall_max)
    penalized_total = round(raw_total * factor, 2)

    # --- Persist penalty info to session if those fields exist ---
    session_changed = False
    if hasattr(session, "penalty_percent"):
        session.penalty_percent = round(overall_max, 2)
        session_changed = True
    if hasattr(session, "penalty_factor"):
        session.penalty_factor = factor
        session_changed = True
    if hasattr(session, "raw_total"):
        session.raw_total = raw_total
        session_changed = True
    if hasattr(session, "penalized_total"):
        session.penalized_total = penalized_total
        session_changed = True
    if hasattr(session, "penalty_applied"):
        session.penalty_applied = (factor < 1.0)
        session_changed = True

    if session_changed:
        # update fields in one save
        update_fields = []
        for f in ("penalty_percent", "penalty_factor", "raw_total", "penalized_total", "penalty_applied"):
            if hasattr(session, f):
                update_fields.append(f)
        # always update end_time too
        if "end_time" not in update_fields and session.end_time:
            update_fields.append("end_time")
        session.save(update_fields=update_fields)

    # Decide which total to show to user: penalized_total (if penalty applied) else raw_total
    display_total = penalized_total if factor < 1.0 else raw_total

    context = {
        'assessment': assessment,
        'submissions': submissions,
        'total_score': display_total,
        'raw_total': raw_total,
        'penalized_total': penalized_total,
        'plag_per_question': per_question_plag,
        'plag_overall_max': round(overall_max, 2),
        'plag_overall_avg': overall_avg,
        'quiz_score': quiz_score,
        'coding_raw_total': round(coding_raw_total, 2),
        'coding_penalized_total': round(coding_penalized_total, 2),
        'penalty_factor': factor,
        'penalty_applied': (factor < 1.0),
    }
    return render(request, 'codingapp/assessment_result.html', context)


from django.views.decorators.http import require_POST
import json

@login_required
@require_POST
def execute_code_api(request, question_id):
    """
    API endpoint to execute code via AJAX, enqueueing the task to Celery
    and returning the Task ID for polling.
    """
    try:
        # 1. Parse the incoming JSON data from the request body
        data = json.loads(request.body)
        code = data.get("code", "").strip()
        language = data.get("language", "python")

        # 2. Check for required data
        if not code or not language:
            return JsonResponse({"success": False, "error": "Missing code or language."}, status=400)

        # 3. Get the question object (needed for ID)
        # Note: We keep this synchronous call as it is fast DB query, and necessary for validation.
        question = get_object_or_404(Question, pk=question_id)
        
        # 4. ⭐ ASYNCHRONOUS CALL: Enqueue the code execution task
        # The task (process_practice_submission) handles running test cases and saving the submission.
        task = process_practice_submission.delay(request.user.id, question.id, code, language)
        
        # 5. Return the Task ID immediately. The client will use this to poll for results.
        return JsonResponse({
            "success": True,
            "task_id": task.id,
            "message": "Code submitted for background processing."
        })

    except json.JSONDecodeError:
        logger.error("JSON Decode Error in execute_code_api")
        return JsonResponse({"success": False, "error": "Invalid JSON payload."}, status=400)
    except Exception as e:
        # Catch unexpected errors (like DB errors or misconfigured imports)
        logger.error(f"Error in execute_code_api: {e}", exc_info=True)
        return JsonResponse({"success": False, "error": f"An unexpected server error occurred: {str(e)}"}, status=500)

# NEW FUNCTION: Export Assessment Leaderboard to CSV

# In codingapp/views.py

@login_required
@user_passes_test(is_admin_or_teacher)
def export_assessment_leaderboard_csv(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)

    # 1. Fetch Participants
    participant_ids = (
        assessment.groups
        .prefetch_related('students')
        .values_list('students', flat=True)
        .distinct()
    )
    participants = User.objects.filter(id__in=participant_ids)

    # 2. Get Ordered Questions (for columns)
    assessment_questions = (
        AssessmentQuestion.objects
        .filter(assessment=assessment)
        .select_related('question')
        .order_by('order')
    )
    questions_list = [aq.question for aq in assessment_questions]

    # 3. Map Submissions: {user_id: {question_id: score}}
    all_submissions = AssessmentSubmission.objects.filter(assessment=assessment)

    user_question_scores = {}
    user_extra_metrics = {}  # 👈 NEW (plagiarism / AI / similarity)

    for sub in all_submissions:
        # Scores (existing logic)
        user_question_scores.setdefault(sub.user_id, {})
        user_question_scores[sub.user_id][sub.question_id] = sub.score

        # 👇 NEW: collect max similarity / plagiarism per user
        extra = user_extra_metrics.setdefault(sub.user_id, {
            'plagiarism': 0.0,
            'ai': 0.0,
            'token': 0.0,
            'structural': 0.0,
        })

        extra['plagiarism'] = max(extra['plagiarism'], float(sub.plagiarism_percent or 0))
        extra['ai'] = max(extra['ai'], float(sub.ai_generated_prob or 0))
        extra['token'] = max(extra['token'], float(sub.token_similarity or 0))
        extra['structural'] = max(extra['structural'], float(sub.structural_similarity or 0))

    # 4. Get Quiz Scores
    quiz_scores_dict = {}
    if assessment.quiz:
        quiz_scores = QuizSubmission.objects.filter(
            quiz=assessment.quiz
        ).values('user__id', 'score')
        quiz_scores_dict = {item['user__id']: item['score'] for item in quiz_scores}

    # 5. Get Session Times
    sessions = AssessmentSession.objects.filter(
        assessment=assessment,
        user__in=participants,
        end_time__isnull=False
    )

    session_times_dict = {}
    for session in sessions:
        if session.end_time and session.start_time:
            duration = session.end_time - session.start_time
            total_seconds = int(duration.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            time_str = f"{hours}h {minutes}m {seconds}s"
            session_times_dict[session.user_id] = {
                'time_str': time_str,
                'sort_key': total_seconds
            }
        else:
            session_times_dict[session.user_id] = {
                'time_str': "N/A",
                'sort_key': float('inf')
            }

    # 6. Build Data Rows
    data_rows = []
    for user in participants:
        quiz_score = quiz_scores_dict.get(user.id, 0)

        current_user_scores = user_question_scores.get(user.id, {})
        coding_total = sum(current_user_scores.values())
        total_score = quiz_score + coding_total

        time_info = session_times_dict.get(
            user.id,
            {'time_str': "N/A", 'sort_key': float('inf')}
        )

        # Only include active participants
        if total_score > 0 or time_info['sort_key'] != float('inf'):
            extra = user_extra_metrics.get(user.id, {})

            data_rows.append({
                'username': user.username,
                'quiz_score': quiz_score,
                'question_scores': [
                    current_user_scores.get(q.id, 0) for q in questions_list
                ],
                'coding_total': coding_total,
                'total_score': total_score,
                'time_str': time_info['time_str'],
                'sort_key': time_info['sort_key'],

                # 👇 NEW FIELDS
                'plagiarism': extra.get('plagiarism', 0.0),
                'ai': extra.get('ai', 0.0),
                'token': extra.get('token', 0.0),
                'structural': extra.get('structural', 0.0),
            })

    # 7. Sort by Score then Time
    data_rows.sort(key=lambda x: (-x['total_score'], x['sort_key']))

    # 8. Write CSV
    response = HttpResponse(content_type='text/csv')
    filename = slugify(assessment.title)
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}_leaderboard.csv"'
    )

    writer = csv.writer(response)

    # Header Row
    header = ['Rank', 'Username']
    if assessment.quiz:
        header.append('Quiz Score')

    for i, q in enumerate(questions_list, 1):
        header.append(f'Q{i} - {q.title}')

    # 👇 NEW COLUMNS ADDED (nothing removed)
    header.extend([
        'Total Coding',
        'Grand Total',
        'Plagiarism %',
        'AI Generated %',
        'Token Similarity %',
        'Structural Similarity %',
        'Time Taken'
    ])
    writer.writerow(header)

    # Data Rows
    for rank, entry in enumerate(data_rows, 1):
        row = [rank, entry['username']]

        if assessment.quiz:
            row.append(entry['quiz_score'])

        row.extend(entry['question_scores'])

        row.extend([
            entry['coding_total'],
            entry['total_score'],
            round(entry['plagiarism'], 2),
            round(entry['ai'] * 100, 2),
            round(entry['token'] * 100, 2),
            round(entry['structural'] * 100, 2),
            entry['time_str'],
        ])

        writer.writerow(row)

    return response


# 👇 ADD THIS ENTIRE NEW VIEW FUNCTION
@login_required
def check_submission_status(request, task_id):
    """
    API endpoint for checking the status of a Celery task (used for polling).
    """
    task = AsyncResult(task_id)
    response_data = {
        'task_id': task_id,
        'status': task.status,
        'ready': task.ready(),
    }
    
    if task.ready():
        try:
            result = task.get(timeout=1) 
            response_data['final_status'] = result.get('final_status')
            response_data['results'] = result.get('results')
            response_data['error'] = result.get('error')
        except Exception as e:
            response_data['final_status'] = 'FAILURE'
            response_data['error'] = str(e)
            
    return JsonResponse(response_data)


# ==============================================================
# 🧩 Permissions Management View (Admin Dashboard)
# ==============================================================
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from codingapp.models import ActionPermission, Role, UserProfile
from codingapp.permissions import can_assign

@login_required
def permissions_manage(request):
    """
    Admin dashboard page to view and assign permissions.
    - Left: list of users
    - Right: selected user's permissions (inherited role perms vs custom perms)
    Supports bulk assignment of multiple custom permissions to a single user.
    """
    from codingapp.models import Role, ActionPermission, UserProfile
    import json

    profile = getattr(request.user, "userprofile", None)
    if not profile or not profile.role or profile.role.name.lower() != "admin":
        messages.error(request, "You don't have access to this page.")
        return redirect("dashboard")

    # Load everything needed
    permissions = ActionPermission.objects.all().order_by("code")
    roles = Role.objects.prefetch_related("permissions").order_by("name")
    users = (
        UserProfile.objects.select_related("role", "user")
        .prefetch_related("custom_permissions", "role__permissions")
        .order_by("role__name", "user__username")   # ✅ Added sorting by role name first
    )
    search_query = request.GET.get("search", "").strip()
    if search_query:
        users = users.filter(
            Q(user__username__icontains=search_query)
            | Q(full_name__icontains=search_query)
            | Q(user__email__icontains=search_query)
            | Q(role__name__icontains=search_query)
        )

    # Build fast lookup maps for template
    role_permission_map = {}
    custom_permission_map = {}
    combined_permission_map = {}

    for u in users:
        role_perms = list(u.role.permissions.values_list("id", flat=True)) if u.role else []
        custom_perms = list(u.custom_permissions.values_list("id", flat=True))
        role_permission_map[u.id] = role_perms
        custom_permission_map[u.id] = custom_perms
        combined_permission_map[u.id] = sorted(set(role_perms) | set(custom_perms))
    
    # Handle POST actions
    if request.method == "POST":
        action = request.POST.get("action")

        # ✅ Bulk update: replace a user’s custom permissions
        if action == "update_user_perms":
            user_id = request.POST.get("user_id")
            perm_ids = request.POST.getlist("perm_ids")
            try:
                user_profile = UserProfile.objects.get(id=user_id)
            except UserProfile.DoesNotExist:
                messages.error(request, "User not found.")
                return redirect("permissions_manage")

            try:
                perm_ids_int = [int(i) for i in perm_ids]
            except ValueError:
                perm_ids_int = []

            perms = ActionPermission.objects.filter(id__in=perm_ids_int)
            user_profile.custom_permissions.set(perms)
            user_profile.save()
            messages.success(request, f"Updated custom permissions for {user_profile.user.username}.")
            return redirect("permissions_manage")

        # ✅ Compatibility: single add/remove actions
        role_id = request.POST.get("role_id")
        perm_id = request.POST.get("perm_id")
        user_id = request.POST.get("user_id")
        single_action = request.POST.get("single_action")

        if role_id and perm_id and single_action:
            try:
                role = Role.objects.get(id=role_id)
                perm = ActionPermission.objects.get(id=perm_id)
                if single_action == "add":
                    role.permissions.add(perm)
                else:
                    role.permissions.remove(perm)
                messages.success(request, f"Permission '{perm.name}' updated for role '{role.name}'.")
            except Exception as e:
                messages.error(request, f"Error: {e}")
            return redirect("permissions_manage")

        if user_id and perm_id and single_action:
            try:
                up = UserProfile.objects.get(id=user_id)
                perm = ActionPermission.objects.get(id=perm_id)
                if single_action == "add":
                    up.custom_permissions.add(perm)
                else:
                    up.custom_permissions.remove(perm)
                messages.success(request, f"Permission '{perm.name}' updated for user '{up.user.username}'.")
            except Exception as e:
                messages.error(request, f"Error: {e}")
            return redirect("permissions_manage")

    # ✅ Serialize maps for template
    context = {
        "permissions": permissions,
        "roles": roles,
        "users": users,
        "role_permission_map": role_permission_map,
        "custom_permission_map": custom_permission_map,
        "user_permission_map": combined_permission_map,
        "role_perm_json": json.dumps(role_permission_map),
        "custom_perm_json": json.dumps(custom_permission_map),
    }
    return render(request, "dashboard/permissions_manage.html", context)

# ==============================================================
# 🧩 HOD Permission Management View
# ==============================================================
from django.db.models import Q
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from codingapp.models import UserProfile, ActionPermission, Role
import json

@login_required
def permissions_hod(request):
    """
    HOD permission management page — similar to admin, but scoped to HOD's department.
    Includes search functionality for easier filtering.
    """
    profile = getattr(request.user, "userprofile", None)
    if not profile or not profile.role or profile.role.name.lower() != "hod":
        messages.error(request, "You don't have access to this page.")
        return redirect("dashboard")

    # Limit scope to users in HOD’s department (excluding self)
    users = (
        UserProfile.objects.filter(
            Q(department=profile.department),
            ~Q(id=profile.id),  # exclude HOD themselves
        )
        .select_related("role", "user", "department")
        .prefetch_related("custom_permissions", "role__permissions")
        .order_by("role__name", "user__username")
    )

    # ✅ Search feature
    search_query = request.GET.get("search", "").strip()
    if search_query:
        users = users.filter(
            Q(user__username__icontains=search_query)
            | Q(full_name__icontains=search_query)
            | Q(user__email__icontains=search_query)
            | Q(role__name__icontains=search_query)
        )

    # Permissions the HOD can manage = those the HOD already has
    allowed_codes = profile.permission_codes()
    permissions = ActionPermission.objects.filter(code__in=allowed_codes).order_by("code")

    # Build maps for frontend JS
    role_permission_map, custom_permission_map, combined_permission_map = {}, {}, {}
    for u in users:
        role_perms = list(u.role.permissions.values_list("id", flat=True)) if u.role else []
        custom_perms = list(u.custom_permissions.values_list("id", flat=True))
        role_permission_map[u.id] = role_perms
        custom_permission_map[u.id] = custom_perms
        combined_permission_map[u.id] = sorted(set(role_perms) | set(custom_perms))

    # ✅ Handle permission updates (HOD can edit users in their dept)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "update_user_perms":
            user_id = request.POST.get("user_id")
            perm_ids = request.POST.getlist("perm_ids")

            try:
                target_user = UserProfile.objects.get(id=user_id, department=profile.department)
            except UserProfile.DoesNotExist:
                messages.error(request, "User not found or not in your department.")
                return redirect("permissions_hod")

            # Only assign permissions that the HOD also possesses
            perms = ActionPermission.objects.filter(id__in=perm_ids, code__in=allowed_codes)
            target_user.custom_permissions.set(perms)
            target_user.save()

            messages.success(
                request,
                f"Updated custom permissions for {target_user.user.username}.",
            )
            return redirect("permissions_hod")

    context = {
        "permissions": permissions,
        "users": users,
        "search_query": search_query,
        "role_perm_json": json.dumps(role_permission_map),
        "custom_perm_json": json.dumps(custom_permission_map),
    }
    return render(request, "dashboard/permissions_hod.html", context)

# ==============================================================
# 🧩 ADMIN CONTROL CENTER
# ==============================================================
from django.db.models import Count

from django.db.models import Count, Q, Prefetch
from django.contrib import messages
from django.shortcuts import render, redirect
from codingapp.models import Department, Role, ActionPermission, UserProfile


@login_required
def admin_control_center(request):
    """Admin's master control center with safe null handling."""
    profile = getattr(request.user, "userprofile", None)
    if not profile or not profile.role or profile.role.name.lower() != "admin":
        messages.error(request, "You don’t have access to this page.")
        return redirect("dashboard")

    # Prefetch safely (avoids N+1)
    roles = Role.objects.prefetch_related("permissions").order_by("name")
    permissions = ActionPermission.objects.all().order_by("code")

    # Annotate departments with member count and prefetch related HOD safely
    departments = (
        Department.objects.select_related("hod__user")
        .annotate(user_count=Count("userprofile"))
        .order_by("name")
    )

    users = (
        UserProfile.objects.select_related("role", "department", "user")
        .order_by("user__username")
    )

    # Count users by role
    role_counts = (
        UserProfile.objects.values("role__name")
        .annotate(count=Count("id"))
        .order_by("role__name")
    )

    context = {
        "roles": roles,
        "permissions": permissions,
        "departments": departments,
        "users": users,
        "role_counts": role_counts,
    }
    return render(request, "dashboard/admin_control_center.html", context)

# ==============================================================
# 🧩 ADMIN: MANAGE USERS
# ==============================================================
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from codingapp.models import Role, Department, UserProfile


@login_required
def admin_manage_users(request):
    """Admin view to manage users, grouped by department and role."""

    profile = getattr(request.user, "userprofile", None)
    if not profile or not profile.role or profile.role.name.lower() != "admin":
        messages.error(request, "Access denied.")
        return redirect("dashboard")

    # Load roles and departments
    from django.db.models.functions import Lower
    roles = (
    Role.objects.annotate(name_lower=Lower("name"))
    .order_by("name_lower")
    .distinct("name_lower")
)
    departments = Department.objects.all().order_by("name")

    # Prefetch all users with relations
    users = (
        UserProfile.objects.select_related("user", "role", "department")
        .prefetch_related("custom_permissions")
        .order_by("department__name", "role__name", "user__username")
    )

    search_query = request.GET.get("search", "").strip()

    users = (
        UserProfile.objects.select_related("user", "role", "department")
        .prefetch_related("custom_permissions")
        .order_by("department__name", "role__name", "user__username")
    )

    # ✅ Apply search filter
    if search_query:
        users = users.filter(
            Q(user__username__icontains=search_query)
            | Q(full_name__icontains=search_query)
            | Q(user__email__icontains=search_query)
            | Q(role__name__icontains=search_query)
        )


    # Group by department and role
    structured_data = []
    for dept in departments:
        dept_users = [u for u in users if u.department == dept]
        if not dept_users:
            continue

        dept_data = {"department": dept, "roles": []}
        for role_name in ["HOD", "Teacher", "Student"]:
            role_users = [
                u for u in dept_users
                if (u.role and u.role.name.lower() == role_name.lower())
            ]
            if role_users:
                dept_data["roles"].append({
                    "role_name": role_name,
                    "users": role_users
                })

        # Add unassigned users within department
        unassigned = [u for u in dept_users if not u.role]
        if unassigned:
            dept_data["roles"].append({
                "role_name": "Unassigned",
                "users": unassigned
            })

        structured_data.append(dept_data)

    # Users without department
    no_dept_users = [u for u in users if not u.department]
    if no_dept_users:
        structured_data.append({
            "department": None,
            "roles": [{"role_name": "Unassigned", "users": no_dept_users}]
        })

    # Handle form actions
    if request.method == "POST":
        action = request.POST.get("action")

        # ADD USER
        if action == "add_user":
            username = request.POST.get("username")
            email = request.POST.get("email")
            password = request.POST.get("password")
            role_id = request.POST.get("role_id")
            dept_id = request.POST.get("dept_id")

            if not username or not password:
                messages.error(request, "Username and password are required.")
                return redirect("admin_manage_users")

            if User.objects.filter(username=username).exists():
                messages.warning(request, f"User '{username}' already exists.")
                return redirect("admin_manage_users")

            user = User.objects.create_user(username=username, email=email, password=password)
            profile = UserProfile.objects.get(user=user)
            if role_id:
                profile.role = Role.objects.get(id=role_id)
            if dept_id:
                profile.department = Department.objects.get(id=dept_id)
            profile.save()

            messages.success(request, f"User '{username}' added successfully.")
            return redirect("admin_manage_users")

        # INDIVIDUAL UPDATE
        elif action == "update_user":
            user_id = request.POST.get("user_id")
            role_id = request.POST.get("role_id")
            dept_id = request.POST.get("dept_id")
            try:
                profile = UserProfile.objects.get(id=user_id)
                if role_id:
                    profile.role = Role.objects.get(id=role_id)
                if dept_id:
                    profile.department = Department.objects.get(id=dept_id)
                profile.save()
                messages.success(request, f"Updated user '{profile.user.username}'.")
            except Exception as e:
                messages.error(request, f"Error updating user: {e}")
            return redirect("admin_manage_users")

        # BULK UPDATE
        elif action == "bulk_update":
            selected_ids = request.POST.getlist("selected_users")
            role_id = request.POST.get("bulk_role")
            dept_id = request.POST.get("bulk_dept")
            if not selected_ids:
                messages.warning(request, "No users selected.")
                return redirect("admin_manage_users")

            count = 0
            for uid in selected_ids:
                try:
                    up = UserProfile.objects.get(id=uid)
                    if role_id:
                        up.role = Role.objects.get(id=role_id)
                    if dept_id:
                        up.department = Department.objects.get(id=dept_id)
                    up.save()
                    count += 1
                except Exception as e:
                    messages.error(request, f"Error updating user {uid}: {e}")
            messages.success(request, f"Bulk updated {count} users successfully.")
            return redirect("admin_manage_users")

        # TOGGLE ACTIVE
        elif action == "toggle_active":
            user_id = request.POST.get("user_id")
            try:
                profile = UserProfile.objects.get(id=user_id)
                profile.user.is_active = not profile.user.is_active
                profile.user.save()
                state = "activated" if profile.user.is_active else "deactivated"
                messages.success(request, f"User '{profile.user.username}' has been {state}.")
            except Exception as e:
                messages.error(request, f"Error toggling user: {e}")
            return redirect("admin_manage_users")

        # DELETE USER
        elif action == "delete_user":
            user_id = request.POST.get("user_id")
            try:
                profile = UserProfile.objects.get(id=user_id)
                username = profile.user.username
                profile.user.delete()
                messages.success(request, f"User '{username}' deleted successfully.")
            except Exception as e:
                messages.error(request, f"Error deleting user: {e}")
            return redirect("admin_manage_users")

        # RESET PASSWORD
        elif action == "reset_password":
            user_id = request.POST.get("user_id")
            new_password = request.POST.get("new_password")
            try:
                profile = UserProfile.objects.get(id=user_id)
                profile.user.set_password(new_password)
                profile.user.save()
                messages.success(request, f"Password reset for '{profile.user.username}'.")
            except Exception as e:
                messages.error(request, f"Error resetting password: {e}")
            return redirect("admin_manage_users")

        # BULK ACTIVATE/DEACTIVATE/DELETE
        elif action in ["bulk_activate", "bulk_deactivate", "bulk_delete"]:
            selected_ids = request.POST.getlist("selected_users")
            if not selected_ids:
                messages.warning(request, "No users selected.")
                return redirect("admin_manage_users")

            count = 0
            for uid in selected_ids:
                try:
                    up = UserProfile.objects.get(id=uid)
                    if action == "bulk_delete":
                        up.user.delete()
                    elif action == "bulk_activate":
                        up.user.is_active = True
                        up.user.save()
                    elif action == "bulk_deactivate":
                        up.user.is_active = False
                        up.user.save()
                    count += 1
                except Exception as e:
                    messages.error(request, f"Error processing user {uid}: {e}")

            verb = (
                "deleted" if action == "bulk_delete"
                else "activated" if action == "bulk_activate"
                else "deactivated"
            )
            messages.success(request, f"{count} users {verb} successfully.")
            return redirect("admin_manage_users")

    context = {
        "roles": roles,
        "departments": departments,
        "structured_data": structured_data,
    }
    return render(request, "dashboard/admin_manage_users.html", context)

@login_required
def admin_manage_departments(request):
    """Admin view to manage departments: create, edit, delete, assign HODs, with user stats."""
    from django.db.models import Count, Q
    from codingapp.models import Department, UserProfile

    # 🔒 Access control
    profile = getattr(request.user, "userprofile", None)
    if not profile or not profile.role or profile.role.name.lower() != "admin":
        messages.error(request, "Access denied.")
        return redirect("dashboard")

    # Annotate departments with teacher & student counts
    departments = (
        Department.objects.select_related("hod")
        .annotate(
            teacher_count=Count(
                "userprofile",
                filter=Q(userprofile__role__name__iexact="teacher"),
            ),
            student_count=Count(
                "userprofile",
                filter=Q(userprofile__role__name__iexact="student"),
            ),
        )
        .order_by("name")
    )

    hod_candidates = (
        UserProfile.objects.filter(role__name__iexact="hod")
        .select_related("user")
        .order_by("user__username")
    )

    # ✅ Handle actions
    if request.method == "POST":
        action = request.POST.get("action")

        # ---------- Add Department ----------
        if action == "add_department":
            name = request.POST.get("name")
            code = request.POST.get("code")
            hod_id = request.POST.get("hod_id")

            if not name or not code:
                messages.error(request, "Name and code are required.")
                return redirect("admin_manage_departments")

            dept, created = Department.objects.get_or_create(name=name, code=code)
            if hod_id:
                try:
                    hod_profile = UserProfile.objects.get(id=hod_id)
                    dept.hod = hod_profile
                except UserProfile.DoesNotExist:
                    messages.warning(request, "Invalid HOD selected.")
            dept.save()
            messages.success(
                request,
                f"Department '{name}' {'added' if created else 'updated'} successfully.",
            )
            return redirect("admin_manage_departments")

        # ---------- Update Department ----------
        elif action == "update_department":
            dept_id = request.POST.get("dept_id")
            name = request.POST.get("name")
            code = request.POST.get("code")
            hod_id = request.POST.get("hod_id")

            try:
                dept = Department.objects.get(id=dept_id)
                dept.name = name
                dept.code = code
                dept.hod = UserProfile.objects.get(id=hod_id) if hod_id else None
                dept.save()
                messages.success(request, f"Updated department '{dept.name}'.")
            except Exception as e:
                messages.error(request, f"Error updating department: {e}")
            return redirect("admin_manage_departments")

        # ---------- Delete Department ----------
        elif action == "delete_department":
            dept_id = request.POST.get("dept_id")
            try:
                dept = Department.objects.get(id=dept_id)
                name = dept.name
                dept.delete()
                messages.success(request, f"Department '{name}' deleted.")
            except Exception as e:
                messages.error(request, f"Error deleting department: {e}")
            return redirect("admin_manage_departments")

    # ---------- Context ----------
    context = {
        "departments": departments,
        "hod_candidates": hod_candidates,
    }
    return render(request, "dashboard/admin_manage_departments.html", context)

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from codingapp.models import Group, Department, UserProfile

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from codingapp.models import Group, Department, UserProfile

@login_required
def manage_groups(request):
    """
    Manage Groups (Admin/HOD full access, Teachers limited to their assigned groups)
    - Admin & HOD: can create/edit/delete groups, assign teachers & students
    - Teachers: can only view and edit groups assigned to them
    """

    profile = getattr(request.user, "userprofile", None)
    if not profile or not profile.role:
        messages.error(request, "Access denied. Missing role information.")
        return redirect("dashboard")

    role_name = profile.role.name.lower()

    # Admin and HOD: full control
    groups = get_user_accessible_groups(request.user).select_related("department").prefetch_related("students", "teachers")

    if role_name in ["admin", "hod"]:
        departments = Department.objects.all().order_by("name")
        teachers = UserProfile.objects.filter(role__name__iexact="teacher").select_related("user")
        students = UserProfile.objects.filter(role__name__iexact="student").select_related("user")

        if role_name == "hod":
            groups = groups.filter(department=profile.department)
            departments = Department.objects.filter(id=profile.department_id)
            teachers = teachers.filter(department=profile.department)
            students = students.filter(department=profile.department)

    elif role_name == "teacher":
        departments = Department.objects.filter(id=profile.department_id)
        teachers = UserProfile.objects.filter(user=request.user)
        students = UserProfile.objects.filter(department=profile.department)

    else:
        messages.error(request, "Access denied.")
        return redirect("dashboard")

    # =======================
    # 🔧 Handle POST Actions
    # =======================
    if request.method == "POST":
        action = request.POST.get("action")

        # ✅ Create or Edit Group (Admin/HOD only)
        if action == "add_or_edit_group" and role_name in ["admin", "hod"]:
            group_id = request.POST.get("group_id")
            name = request.POST.get("name", "").strip()
            dept_id = request.POST.get("department_id")
            teacher_ids = request.POST.getlist("teacher_ids")
            student_ids = request.POST.getlist("student_ids")

            if not name:
                messages.error(request, "Group name cannot be empty.")
                return redirect("manage_groups")

            department = None
            if dept_id:
                department = get_object_or_404(Department, id=dept_id)

            if group_id:
                # Edit
                group = get_object_or_404(Group, id=group_id)
                if role_name == "hod" and group.department != profile.department:
                    messages.error(request, "You can only manage groups in your department.")
                    return redirect("manage_groups")

                group.name = name
                group.department = department
                group.save()
                group.teachers.set([UserProfile.objects.get(id=t).user for t in teacher_ids])
                group.students.set([UserProfile.objects.get(id=s).user for s in student_ids])
                messages.success(request, f"Group '{group.name}' updated successfully.")
            else:
                # Create
                group = Group.objects.create(
                    name=name,
                    department=department,
                    created_by=request.user
                )
                group.teachers.set([UserProfile.objects.get(id=t).user for t in teacher_ids])
                group.students.set([UserProfile.objects.get(id=s).user for s in student_ids])
                messages.success(request, f"Group '{group.name}' created successfully.")

            return redirect("manage_groups")

        # ✅ Delete Group (Admin/HOD only)
        elif action == "delete_group" and role_name in ["admin", "hod"]:
            group_id = request.POST.get("group_id")
            group = get_object_or_404(Group, id=group_id)
            if role_name == "hod" and group.department != profile.department:
                messages.error(request, "You cannot delete groups outside your department.")
            else:
                name = group.name
                group.delete()
                messages.success(request, f"Group '{name}' deleted successfully.")
            return redirect("manage_groups")
    
    if request.method == "POST" and request.POST.get("action") == "teacher_bulk_upload_students":
        profile = getattr(request.user, "userprofile", None)
        group_id = request.POST.get("group_id")
        excel_file = request.FILES.get("excel_file")

        if not profile or not profile.role or profile.role.name.lower() != "teacher":
            messages.error(request, "Access denied. Only teachers can upload students.")
            return redirect("manage_groups")

        if not group_id or not excel_file:
            messages.error(request, "Missing group or file.")
            return redirect("manage_groups")

        # Ensure teacher has access to this group
        group = get_object_or_404(Group, id=group_id)
        if not group.teachers.filter(id=request.user.id).exists():
            messages.error(request, "You can only upload students for your assigned groups.")
            return redirect("manage_groups")

        # ✅ Parse Excel file
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

            required = ["username", "email", "full_name", "password"]
            if not all(h in headers for h in required):
                messages.error(request, f"Missing required headers. Expected: {', '.join(required)}")
                return redirect("manage_groups")

            created = skipped = 0
            dept = profile.department  # auto-assign teacher's department

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))
                username = (data.get("username") or "").strip()
                email = (data.get("email") or "").strip()
                password = (data.get("password") or "password123").strip()
                full_name = (data.get("full_name") or "").strip()

                if not username or not email:
                    continue

                if User.objects.filter(username=username).exists():
                    skipped += 1
                    continue

                user = User.objects.create_user(username=username, email=email, password=password)
                profile_obj, _ = UserProfile.objects.get_or_create(user=user)
                profile_obj.full_name = full_name
                profile_obj.department = dept
                profile_obj.role = Role.objects.filter(name__iexact="student").first()
                profile_obj.save()

                # ✅ Add to this group
                group.students.add(user)
                created += 1

            messages.success(request, f"✅ Uploaded {created} students. {skipped} skipped (existing usernames).")

        except Exception as e:
            messages.error(request, f"Error processing file: {e}")

        return redirect("manage_groups")

    # Prepare context
    context = {
        "groups": groups,
        "departments": departments,
        "teachers": teachers,
        "students": students,
        "role_name": role_name,
    }

    # ✅ Group the groups by department
    from collections import defaultdict
    grouped_by_department = defaultdict(list)
    for g in groups:
        dept_name = g.department.name if g.department else "Unassigned Department"
        grouped_by_department[dept_name].append(g)

    context["grouped_by_department"] = dict(grouped_by_department)

    return render(request, "dashboard/manage_groups.html", context)

@login_required
def external_profile(request):
    from django.contrib import messages
    from django.shortcuts import render, redirect
    from codingapp.forms import ExternalProfileForm
    from codingapp.tasks import sync_external_profiles
    from codingapp.models import ExternalProfile

    # ✅ SAFE: always ensure profile exists
    profile, created = ExternalProfile.objects.get_or_create(
        user=request.user
    )

    if request.method == "POST":
        print("POST DATA:", request.POST)
        form = ExternalProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            sync_external_profiles.delay(request.user.id)
            messages.success(
                request,
                "External profiles saved. Syncing started."
            )
            return redirect("external_profile")
    else:
        form = ExternalProfileForm(instance=profile)

    return render(
        request,
        "codingapp/external_profile.html",
        {
            "form": form,
            "profile": profile
        }
    )

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import json

@login_required
@csrf_exempt
def save_hackerrank_badges(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=400)

    data = json.loads(request.body)
    badges = data.get("badges", [])

    profile = request.user.externalprofile
    profile.hackerrank_badges = badges
    profile.save()

    return JsonResponse({"status": "success", "badges_saved": len(badges)})
